"""Standalone desktop GUI for the personal PV detail analysis system."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import queue
import subprocess
import sys
import threading
import tkinter as tk
from tkinter import messagebox, ttk

from .pv_tracker import ManualCandidate, MemberRecord, PVTracker, format_record_label
from .paths import ANALYSIS_DIR, AUTOCOMMENT_DIR, BACKDESIGN_PATH, PROJECT_ROOT, SRC_DIR, ensure_dirs


BACKGROUND_IMAGE = BACKDESIGN_PATH
MAKE_CARD_SCRIPT = SRC_DIR / "make_card.py"


@dataclass(frozen=True)
class UIConfig:
    title: str = "詳細分析"
    font_family: str = "Segoe UI"
    title_font_size: int = 26
    normal_font_size: int = 11
    background: str = "#F8EEF5"
    panel_background: str = "#FFF8FC"
    accent: str = "#E8A8C8"
    accent_hover: str = "#D98CB3"
    text_color: str = "#4A3B45"
    error_background: str = "#FFF4F8"
    blue_panel_background: str = "#DAF7F6"
    list_background: str = "#FFFFFF"
    background_image_path: str = str(BACKGROUND_IMAGE)


class PVDetailApp(tk.Tk):
    """Tkinter app used locally; it is not connected to the Discord bot."""

    def __init__(self, tracker: PVTracker | None = None, config: UIConfig | None = None) -> None:
        super().__init__()
        self.config_data = config or UIConfig()
        self.tracker = tracker or PVTracker()
        self.worker_queue: queue.Queue[tuple[str, object]] = queue.Queue()
        self.current_result = None
        self.old_record_by_list_label: dict[str, MemberRecord] = {}
        self.new_record_by_list_label: dict[str, MemberRecord] = {}
        self.transfer_by_list_label: dict[str, ManualCandidate] = {}
        self.completed_guild_vars: dict[str, tk.BooleanVar] = {}
        self.pending_lost: list[MemberRecord] = []
        self.manual_link_count = 0
        self.last_moved_lost_count = 0
        self.last_created_new_count = 0
        self.make_card_running = False
        self._build_window()
        self._build_widgets()
        self.refresh_inputs()
        self.after(150, self._drain_worker_queue)

    def _build_window(self) -> None:
        self.title(self.config_data.title)
        self.geometry("1180x760")
        self.minsize(1000, 680)
        self.configure(bg=self.config_data.background)
        self.background_image = None
        self.background_canvas = tk.Canvas(self, bg=self.config_data.background, highlightthickness=0, bd=0)
        self.background_canvas.place(x=0, y=0, relwidth=1, relheight=1)
        self._install_background_image()
        self.option_add("*Font", (self.config_data.font_family, self.config_data.normal_font_size))
        self._apply_styles()
        self.protocol("WM_DELETE_WINDOW", self.on_close)

    def _install_background_image(self) -> None:
        """Prepare a replaceable background image hook for future visual design updates."""
        image_path = self.config_data.background_image_path
        if not image_path:
            return
        path = Path(image_path)
        if not path.exists():
            return
        self.background_image = tk.PhotoImage(file=str(path))
        self.background_canvas.create_image(0, 0, image=self.background_image, anchor="nw")
        # The canvas is created before the UI content window, so no z-order call is needed.


    def _apply_styles(self) -> None:
        style = ttk.Style(self)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass
        font = (self.config_data.font_family, self.config_data.normal_font_size)
        style.configure(".", font=font, foreground=self.config_data.text_color)
        style.configure("Glass.TFrame", background=self.config_data.background)
        style.configure("Panel.TFrame", background=self.config_data.panel_background)
        style.configure(
            "Glass.TLabelframe",
            background=self.config_data.panel_background,
            bordercolor="#E8A8C8",
            relief="solid",
            borderwidth=1,
        )
        style.configure(
            "Glass.TLabelframe.Label",
            background=self.config_data.panel_background,
            foreground=self.config_data.text_color,
            font=(self.config_data.font_family, 12, "bold"),
        )
        style.configure("TLabel", background=self.config_data.panel_background, foreground=self.config_data.text_color)
        style.configure("TCheckbutton", background=self.config_data.panel_background, foreground=self.config_data.text_color)
        style.map("TCheckbutton", background=[("active", self.config_data.panel_background)])
        style.configure(
            "Accent.TButton",
            background=self.config_data.accent,
            foreground=self.config_data.text_color,
            bordercolor="#D98CB3",
            focusthickness=2,
            focuscolor="#D98CB3",
            padding=(14, 8),
            font=(self.config_data.font_family, 10, "bold"),
        )
        style.map("Accent.TButton", background=[("active", self.config_data.accent_hover), ("pressed", "#C977A2")])
        style.configure(
            "Soft.TButton",
            background="#FFFFFF",
            foreground=self.config_data.text_color,
            bordercolor="#D98CB3",
            padding=(12, 8),
            font=(self.config_data.font_family, 10, "bold"),
        )
        style.map("Soft.TButton", background=[("active", "#F5D6E6"), ("pressed", "#E8A8C8")])
        style.configure("TProgressbar", troughcolor="#FFFFFF", background=self.config_data.accent, bordercolor="#E8A8C8")
        style.configure("Vertical.TScrollbar", background="#F5D6E6", troughcolor=self.config_data.panel_background, bordercolor="#E8A8C8", arrowcolor=self.config_data.text_color)

    def _build_widgets(self) -> None:
        root = tk.Canvas(self.background_canvas, bg=self.config_data.background, highlightthickness=0, bd=0)
        if self.background_image is not None:
            root.create_image(0, 0, image=self.background_image, anchor="nw")
        self.content_window = self.background_canvas.create_window(0, 0, anchor="nw", window=root)
        self.background_canvas.bind("<Configure>", self._resize_content_window)
        root.columnconfigure(1, weight=1)
        root.rowconfigure(1, weight=1)

        title = ttk.Label(root, text=self.config_data.title, font=(self.config_data.font_family, self.config_data.title_font_size, "bold"))
        title.grid(row=0, column=0, sticky="w", pady=(0, 10))

        controls = ttk.LabelFrame(root, text="日付選択", padding=10, style="Glass.TLabelframe")
        controls.grid(row=0, column=1, columnspan=2, sticky="ew", padx=(0, 0))
        controls.columnconfigure(6, weight=1)
        ttk.Label(controls, text="旧データ").grid(row=0, column=0, padx=4)
        self.old_date_var = tk.StringVar()
        self.old_combo = ttk.Combobox(controls, textvariable=self.old_date_var, width=14, state="readonly")
        self.old_combo.grid(row=0, column=1, padx=4)
        ttk.Label(controls, text="新データ").grid(row=0, column=2, padx=4)
        self.new_date_var = tk.StringVar()
        self.new_combo = ttk.Combobox(controls, textvariable=self.new_date_var, width=14, state="readonly")
        self.new_combo.grid(row=0, column=3, padx=4)
        ttk.Button(controls, text="Start", command=self.start_analysis, style="Accent.TButton").grid(row=0, column=4, padx=10)
        ttk.Button(controls, text="更新", command=self.refresh_inputs, style="Soft.TButton").grid(row=0, column=5, padx=4)
        ttk.Button(controls, text="原文作成", command=self.create_comment_materials, style="Accent.TButton").grid(row=0, column=7, padx=4)
        ttk.Button(controls, text="ギルドカルテ作成", command=self.run_make_card, style="Accent.TButton").grid(row=0, column=8, padx=(4, 0))

        left = ttk.LabelFrame(root, text="ギルドチェックリスト", padding=10, style="Glass.TLabelframe")
        left.grid(row=1, column=0, sticky="nsew", padx=(0, 10))
        left.rowconfigure(0, weight=1)
        self.guild_canvas = tk.Canvas(left, width=210, highlightthickness=0, bg=self.config_data.panel_background, bd=0)
        self.guild_canvas.grid(row=0, column=0, sticky="nsew")
        guild_scroll = ttk.Scrollbar(left, orient=tk.VERTICAL, command=self.guild_canvas.yview, style="Vertical.TScrollbar")
        guild_scroll.grid(row=0, column=1, sticky="ns")
        self.guild_canvas.configure(yscrollcommand=guild_scroll.set)
        self.guild_frame = tk.Frame(self.guild_canvas, bg=self.config_data.panel_background)
        self.guild_canvas.create_window((0, 0), window=self.guild_frame, anchor="nw")
        self.guild_frame.bind("<Configure>", lambda _event: self.guild_canvas.configure(scrollregion=self.guild_canvas.bbox("all")))

        center = tk.Frame(root, bg=self.config_data.background)
        center.grid(row=1, column=1, sticky="nsew")
        center.columnconfigure(0, weight=1)
        center.rowconfigure(2, weight=1)

        status = ttk.LabelFrame(center, text="処理状況", padding=10, style="Glass.TLabelframe")
        status.grid(row=0, column=0, sticky="ew")
        status.columnconfigure(1, weight=1)
        ttk.Label(status, text="現在のギルド").grid(row=0, column=0, sticky="w")
        self.current_guild_var = tk.StringVar(value="-")
        ttk.Label(status, textvariable=self.current_guild_var, font=(self.config_data.font_family, 18, "bold")).grid(row=0, column=1, sticky="w", padx=12)
        ttk.Label(status, text="現在の家門名").grid(row=1, column=0, sticky="w")
        self.current_family_var = tk.StringVar(value="-")
        ttk.Label(status, textvariable=self.current_family_var, font=(self.config_data.font_family, 16)).grid(row=1, column=1, sticky="w", padx=12)
        self.progress_var = tk.IntVar(value=0)
        self.progress = ttk.Progressbar(status, variable=self.progress_var, maximum=100)
        self.progress.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(10, 0))

        errors = ttk.LabelFrame(center, text="Error / 手動確認", padding=10, style="Glass.TLabelframe")
        errors.grid(row=2, column=0, sticky="nsew", pady=(10, 0))
        errors.columnconfigure(0, weight=1)
        errors.columnconfigure(1, weight=1)
        errors.rowconfigure(1, weight=1)
        errors.rowconfigure(3, weight=1)

        ttk.Label(errors, text="名前不一致：旧データ側").grid(row=0, column=0, sticky="w")
        ttk.Label(errors, text="名前不一致：新データ側").grid(row=0, column=1, sticky="w")
        self.old_list = self._make_listbox(errors, height=9)
        self.old_list.grid(row=1, column=0, sticky="nsew", padx=(0, 5))
        self.new_list = self._make_listbox(errors, height=9)
        self.new_list.grid(row=1, column=1, sticky="nsew", padx=(5, 0))
        self.old_list.bind("<<ListboxSelect>>", lambda _event: self._update_selection_labels())
        self.new_list.bind("<<ListboxSelect>>", lambda _event: self._update_selection_labels())

        ttk.Label(errors, text="移籍候補").grid(row=2, column=0, columnspan=2, sticky="w", pady=(10, 0))
        self.transfer_list = self._make_listbox(errors, height=7)
        self.transfer_list.grid(row=3, column=0, columnspan=2, sticky="nsew")
        self.transfer_list.bind("<<ListboxSelect>>", lambda _event: self._select_transfer_candidate())

        selection = tk.Frame(errors, bg=self.config_data.panel_background)
        selection.grid(row=4, column=0, columnspan=2, sticky="ew", pady=(10, 0))
        selection.columnconfigure(1, weight=1)
        selection.columnconfigure(3, weight=1)
        ttk.Label(selection, text="旧側選択").grid(row=0, column=0, sticky="w")
        self.selected_old_var = tk.StringVar(value="-")
        ttk.Label(selection, textvariable=self.selected_old_var).grid(row=0, column=1, sticky="w")
        ttk.Label(selection, text="新側選択").grid(row=1, column=0, sticky="w")
        self.selected_new_var = tk.StringVar(value="-")
        ttk.Label(selection, textvariable=self.selected_new_var).grid(row=1, column=1, sticky="w")
        ttk.Button(selection, text="✅ 確定", command=self.confirm_manual_link, style="Accent.TButton").grid(row=0, column=4, rowspan=2, padx=12)
        ttk.Button(selection, text="追跡不明へ移動", command=self.move_lost_tracks, style="Accent.TButton").grid(row=0, column=5, rowspan=2, padx=4)
        ttk.Button(selection, text="おわり", command=self.on_close, style="Soft.TButton").grid(row=0, column=6, rowspan=2, padx=4)

        right = ttk.LabelFrame(root, text="メッセージ", padding=10, style="Glass.TLabelframe")
        right.grid(row=1, column=2, sticky="nsew", padx=(10, 0))
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)
        self.log_text = tk.Text(right, width=34, wrap=tk.WORD, bg=self.config_data.blue_panel_background, fg=self.config_data.text_color, relief=tk.FLAT, bd=1, highlightthickness=1, highlightbackground="#E8A8C8", insertbackground=self.config_data.text_color)
        self.log_text.grid(row=0, column=0, sticky="nsew")
        log_scroll = ttk.Scrollbar(right, orient=tk.VERTICAL, command=self.log_text.yview, style="Vertical.TScrollbar")
        log_scroll.grid(row=0, column=1, sticky="ns")
        self.log_text.configure(yscrollcommand=log_scroll.set)


    def _resize_content_window(self, event) -> None:
        self.background_canvas.itemconfigure(self.content_window, width=event.width, height=event.height)

    def _make_listbox(self, parent, height: int) -> tk.Listbox:
        return tk.Listbox(
            parent,
            height=height,
            exportselection=False,
            bg=self.config_data.list_background,
            fg=self.config_data.text_color,
            selectbackground=self.config_data.accent_hover,
            selectforeground="#FFFFFF",
            relief=tk.FLAT,
            bd=1,
            highlightthickness=1,
            highlightbackground="#E8A8C8",
            highlightcolor="#D98CB3",
        )

    def refresh_inputs(self) -> None:
        dates = self.tracker.list_summary_dates()
        self.old_combo["values"] = dates
        self.new_combo["values"] = dates
        if dates:
            self.old_date_var.set(dates[0])
            self.new_date_var.set(dates[-1])
        self._refresh_guild_checklist()
        self.log(f"base: {self.tracker.base_dir}")
        self.log(f"deta_PV: {self.tracker.pv_dir}")

    def _refresh_guild_checklist(self) -> None:
        for child in self.guild_frame.winfo_children():
            child.destroy()
        self.completed_guild_vars.clear()
        for row, guild_name in enumerate(self.tracker.list_guild_names()):
            var = tk.BooleanVar(value=False)
            self.completed_guild_vars[guild_name] = var
            ttk.Checkbutton(self.guild_frame, text=guild_name, variable=var).grid(row=row, column=0, sticky="w", pady=1)

    def start_analysis(self) -> None:
        old_date = self.old_date_var.get()
        new_date = self.new_date_var.get()
        if not old_date or not new_date:
            messagebox.showerror("日付未選択", "旧データと新データを選択してください。")
            return
        self._clear_results()
        self.log(f"分析開始: {old_date} -> {new_date}")
        thread = threading.Thread(target=self._run_analysis_worker, args=(old_date, new_date), daemon=True)
        thread.start()

    def _run_analysis_worker(self, old_date: str, new_date: str) -> None:
        def progress(guild: str, family: str, done: int, total: int) -> None:
            percent = int(done / max(total, 1) * 100)
            self.worker_queue.put(("progress", (guild, family, percent)))
        try:
            result = self.tracker.analyze(old_date, new_date, progress=progress)
        except Exception as exc:  # GUI boundary: show errors without crashing the window.
            self.worker_queue.put(("error", exc))
        else:
            self.worker_queue.put(("done", result))

    def _drain_worker_queue(self) -> None:
        while True:
            try:
                kind, payload = self.worker_queue.get_nowait()
            except queue.Empty:
                break
            if kind == "progress":
                guild, family, percent = payload
                self.current_guild_var.set(guild)
                self.current_family_var.set(family)
                self.progress_var.set(percent)
            elif kind == "error":
                self.log(f"ERROR: {payload}")
                messagebox.showerror("分析エラー", str(payload))
            elif kind == "done":
                self._show_result(payload)
            elif kind == "log":
                self.log(str(payload))
            elif kind == "make_card_done":
                self.make_card_running = False
                self.log(str(payload))
        self.after(150, self._drain_worker_queue)

    def _show_result(self, result) -> None:
        self.current_result = result
        self.pending_lost = list(result.lost_candidates) + list(result.name_mismatches_old)
        self.progress_var.set(100)
        for guild_name in result.processed_guilds:
            if guild_name in self.completed_guild_vars:
                self.completed_guild_vars[guild_name].set(True)
        for record in result.name_mismatches_old + result.lost_candidates:
            label = format_record_label(record)
            self.old_record_by_list_label[label] = record
            self.old_list.insert(tk.END, label)
        for record in result.name_mismatches_new:
            label = format_record_label(record)
            self.new_record_by_list_label[label] = record
            self.new_list.insert(tk.END, label)
        for candidate in result.transfer_candidates:
            label = candidate.label
            self.transfer_by_list_label[label] = candidate
            self.transfer_list.insert(tk.END, label)
        self.log(
            "分析完了: "
            f"完全一致 {len(result.exact_matches)} / "
            f"新規 {len(result.new_players)} / "
            f"名前不一致 旧{len(result.name_mismatches_old)} 新{len(result.name_mismatches_new)} / "
            f"移籍候補 {len(result.transfer_candidates)} / "
            f"追跡不明候補 {len(result.lost_candidates)}"
        )

    def _clear_results(self) -> None:
        for listbox in (self.old_list, self.new_list, self.transfer_list):
            listbox.delete(0, tk.END)
        self.old_record_by_list_label.clear()
        self.new_record_by_list_label.clear()
        self.transfer_by_list_label.clear()
        self.selected_old_var.set("-")
        self.selected_new_var.set("-")
        self.pending_lost.clear()
        self.progress_var.set(0)

    def _selected_from_list(self, listbox: tk.Listbox, mapping: dict[str, MemberRecord]) -> MemberRecord | None:
        selection = listbox.curselection()
        if not selection:
            return None
        label = listbox.get(selection[0])
        return mapping.get(label)

    def _update_selection_labels(self) -> None:
        old_record = self._selected_from_list(self.old_list, self.old_record_by_list_label)
        new_record = self._selected_from_list(self.new_list, self.new_record_by_list_label)
        self.selected_old_var.set(format_record_label(old_record) if old_record else "-")
        self.selected_new_var.set(format_record_label(new_record) if new_record else "-")

    def _select_transfer_candidate(self) -> None:
        selection = self.transfer_list.curselection()
        if not selection:
            return
        candidate = self.transfer_by_list_label.get(self.transfer_list.get(selection[0]))
        if candidate:
            self.selected_old_var.set(format_record_label(candidate.old_record))
            self.selected_new_var.set(format_record_label(candidate.new_record))

    def confirm_manual_link(self) -> None:
        old_record = self._selected_from_list(self.old_list, self.old_record_by_list_label)
        new_record = self._selected_from_list(self.new_list, self.new_record_by_list_label)
        link_type = "manual"
        transfer_selection = self.transfer_list.curselection()
        if transfer_selection:
            candidate = self.transfer_by_list_label.get(self.transfer_list.get(transfer_selection[0]))
            if candidate:
                old_record = candidate.old_record
                new_record = candidate.new_record
                link_type = candidate.reason
        if not old_record or not new_record:
            messagebox.showwarning("選択不足", "旧側と新側を選択してください。")
            return
        result = self.tracker.apply_manual_link(old_record, new_record, link_type=link_type)
        self.manual_link_count += 1
        self.pending_lost = [record for record in self.pending_lost if record.identity_key != old_record.identity_key]
        self.log(f"手動リンク確定: {format_record_label(old_record)} -> {format_record_label(new_record)} ({result.link_type})")
        self._remove_selected_rows()

    def _remove_selected_rows(self) -> None:
        for listbox in (self.old_list, self.new_list, self.transfer_list):
            for index in reversed(listbox.curselection()):
                listbox.delete(index)
        self.selected_old_var.set("-")
        self.selected_new_var.set("-")


    def create_comment_materials(self) -> None:
        output_dir = AUTOCOMMENT_DIR / (self.new_date_var.get().replace("-", "").replace("_", "") or datetime.now().strftime("%Y%m%d"))
        output_dir.mkdir(parents=True, exist_ok=True)
        old_records = self._records_from_listbox(self.old_list, self.old_record_by_list_label)
        new_records = self._records_from_listbox(self.new_list, self.new_record_by_list_label)
        transfer_candidates = self._transfer_candidates_from_listbox()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        old_date = self.old_date_var.get() or "old"
        new_date = self.new_date_var.get() or "new"
        output_path = output_dir / f"autocomment_material_{new_date}_{timestamp}.txt"
        lines = [
            "# コメント材料",
            "",
            f"作成日時: {datetime.now().isoformat(timespec='seconds')}",
            f"旧データ: {old_date}",
            f"新データ: {new_date}",
            "",
        ]
        if self.current_result is not None:
            lines.extend([
                "## 自動処理サマリー",
                f"完全一致: {len(self.current_result.exact_matches)}件",
                f"新規作成済み: {len(self.current_result.new_players)}件",
                f"名前不一致(旧): {len(self.current_result.name_mismatches_old)}件",
                f"名前不一致(新): {len(self.current_result.name_mismatches_new)}件",
                f"移籍候補: {len(self.current_result.transfer_candidates)}件",
                f"追跡不明候補: {len(self.current_result.lost_candidates)}件",
                f"PVカルテに追記した人数: {len(self.current_result.exact_matches) + self.manual_link_count}件",
                f"追跡不明へ送った人数: {self.last_moved_lost_count}件",
                f"新規として作成した人数: {len(self.current_result.new_players) + self.last_created_new_count}件",
                "",
                "## 処理対象ギルド一覧",
                *(self.current_result.processed_guilds or ["なし"]),
                "",
                "## 各ギルドの処理状況",
                *[f"{name}: {'完了' if var.get() else '未完了'}" for name, var in self.completed_guild_vars.items()],
                "",
            ])
        old_lines = [format_record_label(record) for record in old_records] or ["なし"]
        new_lines = [format_record_label(record) for record in new_records] or ["なし"]
        transfer_lines = [candidate.label for candidate in transfer_candidates] or ["なし"]
        lines.extend(["## 画面に残っている旧欄（追跡不明候補）", *old_lines])
        lines.extend(["", "## 画面に残っている新欄（新規プレイヤー候補）", *new_lines])
        lines.extend(["", "## 未処理の移籍候補", *transfer_lines])
        lines.extend(["", "## summary側から取れる主要データ", *self._summary_material_lines(new_date)])
        lines.extend([
            "",
            "## AIに作文させるための注意書き",
            "以下はギルドカルテ用コメントを作るための材料です。数字を無理に全部使わず、自然で読みやすいコメントにしてください。事務的すぎず、煽りすぎず、成長傾向・メンバー変動・注目点が伝わる文章にしてください。",
            "",
            "## メモ",
            "このファイルはAIへ直接送信していません。必要に応じて内容を確認・編集してから利用してください。",
        ])
        output_path.write_text("\n".join(lines), encoding="utf-8")
        self.log(f"原文作成材料を出力しました: {output_path}")
        messagebox.showinfo("原文作成", f"txtを出力しました。\n{output_path}")

    def _summary_material_lines(self, date_value: str) -> list[str]:
        from openpyxl import load_workbook

        normalized = date_value.replace("-", "").replace("_", "")
        candidates = [
            ANALYSIS_DIR / f"summary_{date_value}.xlsx",
            ANALYSIS_DIR / f"summary_{normalized}.xlsx",
        ]
        summary_path = next((path for path in candidates if path.exists()), None)
        if summary_path is None:
            return ["summaryファイルが見つかりません。"]
        lines = [f"summaryファイル: {summary_path.name}"]
        try:
            workbook = load_workbook(summary_path, data_only=True, read_only=True)
            for sheet_name in workbook.sheetnames:
                if sheet_name.lower() == "autocomment":
                    continue
                sheet = workbook[sheet_name]
                lines.append(f"### {sheet_name}")
                for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    values = [str(value) for value in row[:10] if value not in (None, "")]
                    if values:
                        lines.append(" / ".join(values))
                    if row_index >= 20:
                        break
            workbook.close()
        except Exception as exc:
            lines.append(f"summary読み取りエラー: {exc}")
        return lines or ["summaryに抽出可能なデータがありません。"]

    def run_make_card(self) -> None:
        if self.make_card_running:
            self.log("ギルドカルテ作成はすでに実行中です。")
            return
        if not MAKE_CARD_SCRIPT.exists():
            self.log(f"make_card.py が見つかりません: {MAKE_CARD_SCRIPT}")
            messagebox.showerror("ギルドカルテ作成", f"make_card.py が見つかりません。\n{MAKE_CARD_SCRIPT}")
            return
        self.make_card_running = True
        self.log(f"[START] ギルドカルテ作成: {sys.executable} {MAKE_CARD_SCRIPT}")
        thread = threading.Thread(target=self._run_make_card_worker, daemon=True)
        thread.start()

    def _run_make_card_worker(self) -> None:
        process = subprocess.Popen(
            [sys.executable, str(MAKE_CARD_SCRIPT)],
            cwd=str(PROJECT_ROOT),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        assert process.stdout is not None
        for line in process.stdout:
            self.worker_queue.put(("log", line.rstrip()))
        exit_code = process.wait()
        if exit_code == 0:
            self.worker_queue.put(("make_card_done", "[OK] ギルドカルテ作成が完了しました。"))
        else:
            self.worker_queue.put(("make_card_done", f"[FAILED] ギルドカルテ作成に失敗しました: 終了コード {exit_code}"))
        # TODO: make_card.py の出力ファイル名仕様が確定したら、同じギルド名・日付・出力種別のPNGを上書きする命名に寄せる。

    def _records_from_listbox(self, listbox: tk.Listbox, mapping: dict[str, MemberRecord]) -> list[MemberRecord]:
        records: list[MemberRecord] = []
        for index in range(listbox.size()):
            record = mapping.get(listbox.get(index))
            if record is not None:
                records.append(record)
        return records

    def _transfer_candidates_from_listbox(self) -> list[ManualCandidate]:
        candidates: list[ManualCandidate] = []
        for index in range(self.transfer_list.size()):
            candidate = self.transfer_by_list_label.get(self.transfer_list.get(index))
            if candidate is not None:
                candidates.append(candidate)
        return candidates

    def _clear_unresolved_lists(self) -> None:
        for listbox in (self.old_list, self.new_list, self.transfer_list):
            listbox.delete(0, tk.END)
        self.old_record_by_list_label.clear()
        self.new_record_by_list_label.clear()
        self.transfer_by_list_label.clear()
        self.pending_lost.clear()
        self.selected_old_var.set("-")
        self.selected_new_var.set("-")

    def move_lost_tracks(self) -> None:
        if not self.current_result:
            return
        old_records = self._records_from_listbox(self.old_list, self.old_record_by_list_label)
        new_records = self._records_from_listbox(self.new_list, self.new_record_by_list_label)
        transfer_candidates = self._transfer_candidates_from_listbox()
        total_targets = len(old_records) + len(new_records) + len(transfer_candidates)
        if total_targets == 0:
            messagebox.showinfo("追跡不明", "移動・作成対象はありません。")
            return
        message = (
            f"旧欄 {len(old_records)}件を追跡不明へ保存/移動し、"
            f"新欄 {len(new_records)}件を新規PVカルテ作成します。\n"
            f"未処理の移籍候補 {len(transfer_candidates)}件は、旧側を追跡不明・新側を新規作成します。\n"
            "実行しますか？"
        )
        if not messagebox.askyesno("未処理分を確定", message):
            return
        moved, created = self.tracker.finalize_unresolved_records(
            old_records,
            new_records,
            transfer_candidates,
            self.current_result.new_date,
            note="GUI unresolved finalize",
        )
        self.last_moved_lost_count += len(moved)
        self.last_created_new_count += len(created)
        self.log(f"未処理分を確定: 追跡不明 {len(moved)}件 / 新規 {len(created)}件")
        self._clear_unresolved_lists()

    def _has_unresolved_items(self) -> bool:
        return any(listbox.size() for listbox in (self.old_list, self.new_list, self.transfer_list))

    def on_close(self) -> None:
        incomplete_guilds = [name for name, var in self.completed_guild_vars.items() if not var.get()]
        if incomplete_guilds or self.pending_lost or self._has_unresolved_items():
            if not messagebox.askyesno("終了確認", "まだ終わっていませんが終了していいですか？"):
                return
        self.destroy()

    def log(self, message: str) -> None:
        self.log_text.insert(tk.END, message + "\n")
        self.log_text.see(tk.END)


def main() -> None:
    ensure_dirs()
    app = PVDetailApp()
    app.mainloop()


if __name__ == "__main__":
    main()
