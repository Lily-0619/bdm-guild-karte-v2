"""Personal PV chart tracking engine for the standalone detail analysis app.

This module is intentionally independent from the Discord bot and from the
existing guild-card application flow.  It reads guild member workbooks, updates
per-person PV workbooks, and stores manual/lost tracking history in SQLite.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
import os
import re
import shutil
import sqlite3
from typing import Callable, Iterable

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from openpyxl import Workbook

try:
    from .paths import ANALYSIS_DIR, DATA_DIR, DB_PATH, DETA_PV_DIR, PROJECT_ROOT, PV_TEMPLATE_PATH, ensure_dirs
except ImportError:  # 直接実行された場合のため
    from paths import ANALYSIS_DIR, DATA_DIR, DB_PATH, DETA_PV_DIR, PROJECT_ROOT, PV_TEMPLATE_PATH, ensure_dirs  # type: ignore

INVALID_FILENAME_CHARS = r'\\/:*?"<>|'
DATE_PATTERNS = (
    re.compile(r"(?P<year>20\d{2})[-_](?P<month>\d{1,2})[-_](?P<day>\d{1,2})"),
    re.compile(r"(?P<year>20\d{2})(?P<month>\d{2})(?P<day>\d{2})"),
)

ProgressCallback = Callable[[str, str, int, int], None]
LOST_STATUS_MARK = "×"


@dataclass(slots=True, frozen=True)
class MemberRecord:
    """One member row from a guild workbook."""

    date: str
    family_name: str
    cpm: float | int | None
    guild_name: str
    source_file: Path | None = None

    @property
    def identity_key(self) -> tuple[str, str]:
        return (self.guild_name, self.family_name)

    @property
    def family_key(self) -> str:
        return self.family_name


@dataclass(slots=True)
class ManualCandidate:
    """A pair or one-sided row requiring human confirmation."""

    old_record: MemberRecord | None
    new_record: MemberRecord | None
    reason: str

    @property
    def label(self) -> str:
        old_label = format_record_label(self.old_record) if self.old_record else "旧なし"
        new_label = format_record_label(self.new_record) if self.new_record else "新なし"
        return f"{old_label}  →  {new_label}"


@dataclass(slots=True)
class AnalysisResult:
    """Summary returned after comparing two dates."""

    old_date: str
    new_date: str
    processed_guilds: list[str] = field(default_factory=list)
    exact_matches: list[tuple[MemberRecord, MemberRecord]] = field(default_factory=list)
    new_players: list[MemberRecord] = field(default_factory=list)
    name_mismatches_old: list[MemberRecord] = field(default_factory=list)
    name_mismatches_new: list[MemberRecord] = field(default_factory=list)
    transfer_candidates: list[ManualCandidate] = field(default_factory=list)
    lost_candidates: list[MemberRecord] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass(slots=True)
class LinkResult:
    """Result of applying a manual link."""

    old_file: Path
    new_file: Path
    appended: bool
    link_type: str


class PVTracker:
    """Read guild data and maintain personal PV chart workbooks."""

    def __init__(self, base_dir: str | Path | None = None) -> None:
        self.base_dir = resolve_base_dir(base_dir)
        self.analysis_dir = ANALYSIS_DIR if self.base_dir == PROJECT_ROOT else self.base_dir / "analysis"
        self.data_dir = DATA_DIR if self.base_dir == PROJECT_ROOT else self.base_dir / "data"
        self.pv_dir = DETA_PV_DIR if self.base_dir == PROJECT_ROOT else self.base_dir / "deta_PV"
        self.template_path = PV_TEMPLATE_PATH if self.base_dir == PROJECT_ROOT else self.base_dir / "template" / "pv_karute.xlsx"
        self.db_path = DB_PATH if self.base_dir == PROJECT_ROOT else self.data_dir / "bdm_guild.sqlite3"
        self.ensure_directories()
        self.ensure_database()

    def ensure_directories(self) -> None:
        if self.base_dir == PROJECT_ROOT:
            ensure_dirs()
        else:
            self.pv_dir.mkdir(parents=True, exist_ok=True)
            self.data_dir.mkdir(parents=True, exist_ok=True)
            self.analysis_dir.mkdir(parents=True, exist_ok=True)

    def ensure_database(self) -> None:
        self.data_dir.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS pv_identity_links (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    old_date TEXT,
                    new_date TEXT,
                    old_family_name TEXT,
                    old_guild_name TEXT,
                    new_family_name TEXT,
                    new_guild_name TEXT,
                    old_pv_file TEXT,
                    new_pv_file TEXT,
                    link_type TEXT,
                    linked_at TEXT,
                    note TEXT
                )
                """
            )
            conn.execute(
                """
                CREATE TABLE IF NOT EXISTS pv_lost_tracks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    lost_date TEXT,
                    family_name TEXT,
                    guild_name TEXT,
                    previous_pv_file TEXT,
                    moved_to TEXT,
                    created_at TEXT,
                    note TEXT
                )
                """
            )

    def storage_path(self, path: Path | None) -> str:
        if path is None:
            return ""
        try:
            return path.resolve().relative_to(self.base_dir).as_posix()
        except ValueError:
            return str(path)

    def list_summary_dates(self) -> list[str]:
        """Return selectable dates discovered from analysis/summary_*.xlsx."""
        if not self.analysis_dir.exists():
            return []
        dates = {
            normalize_date(path.stem.replace("summary_", ""))
            for path in self.analysis_dir.glob("summary_*.xlsx")
        }
        return sorted(date for date in dates if date)

    def list_guild_names(self) -> list[str]:
        if not self.data_dir.exists():
            return []
        return sorted(path.name for path in self.data_dir.iterdir() if path.is_dir())

    def load_members_for_date(self, date_value: str) -> list[MemberRecord]:
        target_date = normalize_date(date_value)
        records: list[MemberRecord] = []
        for workbook_path in self.find_guild_workbooks(target_date):
            records.extend(self.read_members_workbook(workbook_path, target_date))
        return records

    def find_guild_workbooks(self, date_value: str) -> list[Path]:
        target_date = normalize_date(date_value)
        if not self.data_dir.exists():
            return []
        found: list[Path] = []
        for guild_dir in sorted(path for path in self.data_dir.iterdir() if path.is_dir()):
            for workbook_path in sorted(guild_dir.glob("guild_*.xlsx")):
                workbook_date = normalize_date(workbook_path.stem)
                if workbook_date == target_date:
                    found.append(workbook_path)
        return found

    def read_members_workbook(self, workbook_path: Path, date_value: str) -> list[MemberRecord]:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, data_only=True, read_only=True)
        if "members" not in workbook.sheetnames:
            workbook.close()
            return []
        sheet = workbook["members"]
        header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        columns = detect_member_columns(header_row)
        fallback_guild = workbook_path.parent.name
        records: list[MemberRecord] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            family_name = value_at(row, columns.get("family_name"))
            if family_name in (None, ""):
                continue
            guild_name = value_at(row, columns.get("guild_name")) or fallback_guild
            row_date = normalize_date(str(value_at(row, columns.get("date")) or date_value)) or normalize_date(date_value)
            records.append(
                MemberRecord(
                    date=row_date,
                    family_name=str(family_name).strip(),
                    cpm=parse_number(value_at(row, columns.get("cpm"))),
                    guild_name=str(guild_name).strip(),
                    source_file=workbook_path,
                )
            )
        workbook.close()
        return records

    def analyze(self, old_date: str, new_date: str, progress: ProgressCallback | None = None) -> AnalysisResult:
        old_normalized = normalize_date(old_date)
        new_normalized = normalize_date(new_date)
        result = AnalysisResult(old_date=old_normalized, new_date=new_normalized)
        old_records = self.load_members_for_date(old_normalized)
        new_records = self.load_members_for_date(new_normalized)
        old_by_identity = {record.identity_key: record for record in old_records}
        new_by_identity = {record.identity_key: record for record in new_records}
        old_by_family = build_family_index(old_records)
        new_by_family = build_family_index(new_records)
        exact_keys = sorted(set(old_by_identity) & set(new_by_identity))
        total = max(len(exact_keys) + len(set(new_by_identity) - set(old_by_identity)), 1)
        done = 0

        for key in exact_keys:
            old_record = old_by_identity[key]
            new_record = new_by_identity[key]
            done += 1
            if progress:
                progress(new_record.guild_name, new_record.family_name, done, total)
            self.append_existing_or_create(old_record, new_record)
            result.exact_matches.append((old_record, new_record))
            if new_record.guild_name not in result.processed_guilds:
                result.processed_guilds.append(new_record.guild_name)

        unmatched_old_keys = set(old_by_identity) - set(new_by_identity)
        unmatched_new_keys = set(new_by_identity) - set(old_by_identity)
        consumed_old: set[tuple[str, str]] = set()
        consumed_new: set[tuple[str, str]] = set()

        for family_name, old_family_records in old_by_family.items():
            new_family_records = new_by_family.get(family_name, [])
            for old_record in old_family_records:
                for new_record in new_family_records:
                    if old_record.guild_name != new_record.guild_name:
                        result.transfer_candidates.append(ManualCandidate(old_record, new_record, "guild_transfer"))
                        consumed_old.add(old_record.identity_key)
                        consumed_new.add(new_record.identity_key)

        old_unmatched_guilds = {old_by_identity[key].guild_name for key in unmatched_old_keys if key not in consumed_old}
        new_unmatched_guilds = {new_by_identity[key].guild_name for key in unmatched_new_keys if key not in consumed_new}

        for key in sorted(unmatched_new_keys):
            if key in consumed_new:
                continue
            new_record = new_by_identity[key]
            if new_record.guild_name in old_unmatched_guilds:
                result.name_mismatches_new.append(new_record)
            else:
                self.create_or_append_new_player(new_record)
                result.new_players.append(new_record)
            done += 1
            if progress:
                progress(new_record.guild_name, new_record.family_name, done, total)

        for key in sorted(unmatched_old_keys):
            if key in consumed_old:
                continue
            old_record = old_by_identity[key]
            if old_record.guild_name in new_unmatched_guilds:
                result.name_mismatches_old.append(old_record)
            else:
                result.lost_candidates.append(old_record)

        result.processed_guilds = sorted({record.guild_name for record in new_records})
        self.ensure_database()
        return result

    def append_existing_or_create(self, old_record: MemberRecord, new_record: MemberRecord) -> Path:
        workbook_path = self.find_pv_file(old_record)
        if workbook_path is None:
            workbook_path = self.create_person_workbook(old_record, include_initial_row=True)
        self.append_record(workbook_path, new_record)
        final_path = self.rename_pv_file(workbook_path, new_record.guild_name, new_record.family_name)
        return final_path

    def create_or_append_new_player(self, record: MemberRecord) -> Path:
        path = self.find_pv_file(record)
        if path is None:
            path = self.create_person_workbook(record, include_initial_row=True)
        else:
            self.append_record(path, record)
        return self.rename_pv_file(path, record.guild_name, record.family_name)

    def create_person_workbook(self, record: MemberRecord, include_initial_row: bool = False) -> Path:
        target_path = unique_path(self.pv_dir / make_pv_filename(record.guild_name, record.family_name))
        if self.template_path.exists():
            shutil.copy2(self.template_path, target_path)
            from openpyxl import load_workbook

            workbook = load_workbook(target_path)
            ensure_sheets(workbook)
        else:
            workbook = create_blank_pv_workbook()
        if include_initial_row:
            append_record_to_workbook(workbook, record, calculate_growth=False)
        update_charts(workbook)
        workbook.save(target_path)
        workbook.close()
        return target_path

    def append_record(self, workbook_path: Path, record: MemberRecord) -> bool:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path)
        ensure_sheets(workbook)
        appended = append_record_to_workbook(workbook, record, calculate_growth=True)
        update_charts(workbook)
        workbook.save(workbook_path)
        workbook.close()
        return appended

    def find_pv_file(self, record: MemberRecord) -> Path | None:
        preferred = self.pv_dir / make_pv_filename(record.guild_name, record.family_name)
        if preferred.exists():
            return preferred
        suffix = f"_{sanitize_filename_part(record.family_name)}.xlsx"
        matches = sorted(self.pv_dir.glob(f"*{suffix}"))
        return matches[0] if matches else None

    def rename_pv_file(self, current_path: Path, guild_name: str, family_name: str) -> Path:
        target = self.pv_dir / make_pv_filename(guild_name, family_name)
        if current_path.resolve() == target.resolve():
            return current_path
        if target.exists():
            target = unique_path(target)
        current_path.rename(target)
        return target

    def apply_manual_link(
        self,
        old_record: MemberRecord,
        new_record: MemberRecord,
        link_type: str = "manual",
        note: str = "",
    ) -> LinkResult:
        old_file = self.find_pv_file(old_record) or self.create_person_workbook(old_record, include_initial_row=True)
        self.append_record(old_file, new_record)
        new_file = self.rename_pv_file(old_file, new_record.guild_name, new_record.family_name)
        if link_type == "manual":
            if old_record.family_name != new_record.family_name:
                link_type = "name_change"
            elif old_record.guild_name != new_record.guild_name:
                link_type = "guild_transfer"
        self.save_identity_link(old_record, new_record, old_file, new_file, link_type, note)
        return LinkResult(old_file=old_file, new_file=new_file, appended=True, link_type=link_type)

    def save_identity_link(
        self,
        old_record: MemberRecord,
        new_record: MemberRecord,
        old_file: Path,
        new_file: Path,
        link_type: str,
        note: str = "",
    ) -> None:
        self.ensure_database()
        with sqlite3.connect(self.db_path) as conn:
            conn.execute(
                """
                INSERT INTO pv_identity_links (
                    old_date, new_date, old_family_name, old_guild_name,
                    new_family_name, new_guild_name, old_pv_file, new_pv_file,
                    link_type, linked_at, note
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    old_record.date,
                    new_record.date,
                    old_record.family_name,
                    old_record.guild_name,
                    new_record.family_name,
                    new_record.guild_name,
                    self.storage_path(old_file),
                    self.storage_path(new_file),
                    link_type,
                    datetime.now().isoformat(timespec="seconds"),
                    note,
                ),
            )

    def move_lost_tracks(self, records: Iterable[MemberRecord], lost_date: str, note: str = "") -> list[tuple[Path, Path]]:
        target_dir = self.pv_dir / "追跡不明" / normalize_date(lost_date)
        target_dir.mkdir(parents=True, exist_ok=True)
        moved: list[tuple[Path, Path]] = []
        self.ensure_database()
        with sqlite3.connect(self.db_path) as conn:
            for record in unique_records(records):
                source = self.find_pv_file(record)
                previous_file = self.storage_path(source)
                if source is None or not source.exists():
                    destination = self.create_lost_person_workbook(record, target_dir)
                else:
                    self.mark_lost_status(source, record)
                    destination = unique_path(target_dir / source.name)
                    shutil.move(str(source), str(destination))
                conn.execute(
                    """
                    INSERT INTO pv_lost_tracks (
                        lost_date, family_name, guild_name, previous_pv_file,
                        moved_to, created_at, note
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        normalize_date(lost_date),
                        record.family_name,
                        record.guild_name,
                        previous_file,
                        self.storage_path(destination),
                        datetime.now().isoformat(timespec="seconds"),
                        note,
                    ),
                )
                moved.append((source or destination, destination))
        return moved

    def finalize_unresolved_records(
        self,
        old_records: Iterable[MemberRecord],
        new_records: Iterable[MemberRecord],
        transfer_candidates: Iterable[ManualCandidate],
        lost_date: str,
        note: str = "",
    ) -> tuple[list[tuple[Path, Path]], list[Path]]:
        lost_targets = list(old_records)
        new_targets = list(new_records)
        for candidate in transfer_candidates:
            if candidate.old_record is not None:
                lost_targets.append(candidate.old_record)
            if candidate.new_record is not None:
                new_targets.append(candidate.new_record)

        moved_lost = self.move_lost_tracks(unique_records(lost_targets), lost_date, note=note)
        created_new = [self.create_or_append_new_player(record) for record in unique_records(new_targets)]
        return moved_lost, created_new

    def create_lost_person_workbook(self, record: MemberRecord, target_dir: Path) -> Path:
        destination = unique_path(target_dir / make_pv_filename(record.guild_name, record.family_name))
        if self.template_path.exists():
            shutil.copy2(self.template_path, destination)
            from openpyxl import load_workbook

            workbook = load_workbook(destination)
            ensure_sheets(workbook)
        else:
            workbook = create_blank_pv_workbook()
        append_record_to_workbook(workbook, record, calculate_growth=False, status=LOST_STATUS_MARK)
        update_charts(workbook)
        workbook.save(destination)
        workbook.close()
        return destination

    def mark_lost_status(self, workbook_path: Path, record: MemberRecord) -> None:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path)
        ensure_sheets(workbook)
        data_sheet = workbook["データ"]
        target_row = find_matching_data_row(data_sheet, record) or data_sheet.max_row
        if target_row < 2:
            append_record_to_workbook(workbook, record, calculate_growth=False, status=LOST_STATUS_MARK)
        else:
            data_sheet.cell(row=target_row, column=6, value=LOST_STATUS_MARK)
        update_charts(workbook)
        workbook.save(workbook_path)
        workbook.close()


def resolve_base_dir(base_dir: str | Path | None = None) -> Path:
    if base_dir:
        return Path(base_dir).expanduser().resolve()
    env_path = Path(str_path) if (str_path := os.environ.get("BDM_GUILD_KARTE_HOME")) else None
    if env_path:
        return env_path.expanduser().resolve()
    return PROJECT_ROOT


def normalize_date(value: str | None) -> str:
    if not value:
        return ""
    text = str(value)
    for pattern in DATE_PATTERNS:
        match = pattern.search(text)
        if match:
            return f"{int(match.group('year')):04d}{int(match.group('month')):02d}{int(match.group('day')):02d}"
    return ""


def sanitize_filename_part(value: str, max_length: int = 80) -> str:
    text = str(value).strip() or "unknown"
    for char in INVALID_FILENAME_CHARS:
        text = text.replace(char, "_")
    text = re.sub(r"\s+", "_", text).strip(" ._") or "unknown"
    return text[:max_length]


def make_pv_filename(guild_name: str, family_name: str) -> str:
    return f"{sanitize_filename_part(guild_name, 60)}_{sanitize_filename_part(family_name, 100)}.xlsx"


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    counter = 1
    while True:
        candidate = parent / f"{stem}_{counter:02d}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def value_at(row: tuple, index: int | None):
    if index is None or index >= len(row):
        return None
    return row[index]


def parse_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).replace(",", "").strip()
    try:
        number = float(text)
    except ValueError:
        return None
    return int(number) if number.is_integer() else number


def detect_member_columns(header_row: tuple) -> dict[str, int]:
    mapping: dict[str, int] = {}
    aliases = {
        "date": ("日付", "date", "取得日", "集計日"),
        "family_name": ("家門名", "family", "family_name", "name", "名前", "プレイヤー"),
        "cpm": ("cpm", "CPM", "戦闘力", "cp"),
        "guild_name": ("所属", "ギルド", "guild", "guild_name", "所属ギルド"),
    }
    normalized_headers = [str(value).strip().lower() if value is not None else "" for value in header_row]
    for key, names in aliases.items():
        for index, header in enumerate(normalized_headers):
            if any(str(name).lower() in header for name in names):
                mapping[key] = index
                break
    mapping.setdefault("family_name", 0)
    mapping.setdefault("cpm", 1)
    return mapping


def build_family_index(records: Iterable[MemberRecord]) -> dict[str, list[MemberRecord]]:
    index: dict[str, list[MemberRecord]] = {}
    for record in records:
        index.setdefault(record.family_name, []).append(record)
    return index


def unique_records(records: Iterable[MemberRecord]) -> list[MemberRecord]:
    unique: dict[tuple[str, str, str], MemberRecord] = {}
    for record in records:
        unique.setdefault((record.date, record.guild_name, record.family_name), record)
    return list(unique.values())


def create_blank_pv_workbook() -> "Workbook":
    from openpyxl import Workbook

    workbook = Workbook()
    data_sheet = workbook.active
    data_sheet.title = "データ"
    workbook.create_sheet("グラフ")
    ensure_sheets(workbook)
    return workbook


def ensure_sheets(workbook: "Workbook") -> None:
    if "データ" not in workbook.sheetnames:
        workbook.create_sheet("データ", 0)
    if "グラフ" not in workbook.sheetnames:
        workbook.create_sheet("グラフ")
    data_sheet = workbook["データ"]
    headers = ["日付", "CPM", "伸び", "家門名", "所属", "状態"]
    for column, header in enumerate(headers, start=1):
        if data_sheet.cell(row=1, column=column).value in (None, ""):
            data_sheet.cell(row=1, column=column, value=header)


def append_record_to_workbook(
    workbook: "Workbook",
    record: MemberRecord,
    calculate_growth: bool,
    status: str = "",
) -> bool:
    data_sheet = workbook["データ"]
    last_row = data_sheet.max_row
    if last_row >= 2 and str(data_sheet.cell(last_row, 1).value) == record.date:
        if str(data_sheet.cell(last_row, 4).value) == record.family_name and str(data_sheet.cell(last_row, 5).value) == record.guild_name:
            return False
    previous_cpm = None
    if calculate_growth:
        for row in range(data_sheet.max_row, 1, -1):
            previous_cpm = parse_number(data_sheet.cell(row=row, column=2).value)
            if previous_cpm is not None:
                break
    growth = None
    if calculate_growth and previous_cpm is not None and record.cpm is not None:
        growth = record.cpm - previous_cpm
    data_sheet.append([record.date, record.cpm, growth, record.family_name, record.guild_name, status])
    return True


def find_matching_data_row(data_sheet, record: MemberRecord) -> int | None:
    for row in range(data_sheet.max_row, 1, -1):
        row_date = normalize_date(str(data_sheet.cell(row=row, column=1).value or ""))
        family_name = str(data_sheet.cell(row=row, column=4).value or "").strip()
        guild_name = str(data_sheet.cell(row=row, column=5).value or "").strip()
        if row_date == record.date and family_name == record.family_name and guild_name == record.guild_name:
            return row
    return None


def update_charts(workbook: "Workbook") -> None:
    ensure_sheets(workbook)
    from openpyxl.chart import LineChart, Reference

    data_sheet = workbook["データ"]
    graph_sheet = workbook["グラフ"]
    graph_sheet._charts = []
    max_row = data_sheet.max_row
    if max_row < 2:
        return
    categories = Reference(data_sheet, min_col=1, min_row=2, max_row=max_row)

    cpm_chart = LineChart()
    cpm_chart.title = "CPM推移"
    cpm_chart.y_axis.title = "CPM"
    cpm_chart.x_axis.title = "日付"
    cpm_chart.add_data(Reference(data_sheet, min_col=2, min_row=1, max_row=max_row), titles_from_data=True)
    cpm_chart.set_categories(categories)
    graph_sheet.add_chart(cpm_chart, "A1")

    growth_chart = LineChart()
    growth_chart.title = "伸び推移"
    growth_chart.y_axis.title = "伸び"
    growth_chart.x_axis.title = "日付"
    growth_chart.add_data(Reference(data_sheet, min_col=3, min_row=1, max_row=max_row), titles_from_data=True)
    growth_chart.set_categories(categories)
    graph_sheet.add_chart(growth_chart, "A18")


def format_record_label(record: MemberRecord | None) -> str:
    if record is None:
        return ""
    return f"{record.guild_name} / {record.family_name} / CPM:{record.cpm}"
