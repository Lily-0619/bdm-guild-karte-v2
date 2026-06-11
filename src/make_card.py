"""Guild card generator.

Reads the latest analysis summary and per-guild member workbooks, fills the
Excel card template, and writes completed guild cards as XLSX files.  PNG export
is attempted only on Windows with Excel/pywin32 available.
"""

from __future__ import annotations

import importlib
import importlib.util
import logging
import math
import platform
import re
import shutil
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.styles import Border, PatternFill, Side

try:
    from .paths import ANALYSIS_DIR, CARDS_DIR, CONFIG_DIR, DATA_DIR, PROJECT_ROOT, ensure_dirs
except ImportError:  # 直接実行された場合のため
    from paths import ANALYSIS_DIR, CARDS_DIR, CONFIG_DIR, DATA_DIR, PROJECT_ROOT, ensure_dirs  # type: ignore

ROOT_DIR = PROJECT_ROOT
CONFIG_PATH = CONFIG_DIR / "card_guilds.txt"
TEMPLATE_PATH = ROOT_DIR / "template" / "karte.xlsx"
OUTPUT_DIR = CARDS_DIR

SUMMARY_PATTERN = "summary_*.xlsx"
GUILD_PATTERN = "guild_*.xlsx"
TEMPLATE_CARD_SHEET = "カルテ"
TEMPLATE_MEMBERS_SHEET = "個人戦闘力一覧"
MEMBER_LEFT_COLUMNS = (2, 3, 4)
MEMBER_RIGHT_COLUMNS = (6, 7, 8)
MEMBER_HEADER_LABELS = ("No", "家門名", "CPM")
MEMBER_MAX_ROWS_PER_SIDE = 25
MEMBER_MAX_COUNT = MEMBER_MAX_ROWS_PER_SIDE * 2
PINK_FILL = "D98CB3"
LIGHT_PINK_FILL = "FCEFF4"
GRAPH_BORDER = "F0CAD8"
GRAPH_START_ROW = 12
GRAPH_LABEL_COL = 2
GRAPH_COUNT_COL = 3
GRAPH_START_COL = 4
GRAPH_END_COL = 13
GRAPH_WIDTH = GRAPH_END_COL - GRAPH_START_COL + 1
INVALID_FILENAME_CHARS = re.compile(r'[\\/:*?"<>|]')
PLACEHOLDER_PATTERN = re.compile(r"\{\{\s*([A-Za-z0-9_]+)\s*\}\}")
EXACT_PLACEHOLDER_PATTERN = re.compile(r"^\s*\{\{\s*([A-Za-z0-9_]+)\s*\}\}\s*$")
NUMERIC_PLACEHOLDER_KEYS = {"member_count", "rank_by_avg_cpm"}
SUMMARY_DATE_PATTERN = re.compile(r"summary_(\d{4}-\d{2}-\d{2})\.xlsx$", re.IGNORECASE)

CPM_KEYS = {
    "avg_cpm",
    "median_cpm",
    "max_cpm",
    "min_cpm",
    "stdev_cpm",
    "top10_avg_cpm",
    "top15_avg_cpm",
    "top20_avg_cpm",
    "top25_avg_cpm",
    "prev_avg_cpm",
    "avg_cpm_growth",
}
COUNT_KEYS = {
    "member_count",
    "rank_by_avg_cpm",
    "cpm_130000_plus_count",
    "cpm_120000_129999_count",
    "cpm_110000_119999_count",
    "cpm_100000_109999_count",
    "cpm_90000_99999_count",
    "cpm_under_90000_count",
}
RATE_KEYS = {"node_win_rate", "siege_win_rate", "avg_cpm_growth_rate"}
CHART_BANDS = [
    ("130k+", "cpm_130000_plus_count"),
    ("120k-129k", "cpm_120000_129999_count"),
    ("110k-119k", "cpm_110000_119999_count"),
    ("100k-109k", "cpm_100000_109999_count"),
    ("90k-99k", "cpm_90000_99999_count"),
    ("Below 90k", "cpm_under_90000_count"),
]

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)


class CardGenerationError(RuntimeError):
    """Raised for fatal card generation setup errors."""


def sanitize_filename(value: str) -> str:
    """Replace only Windows-forbidden filename characters with underscores."""
    sanitized = INVALID_FILENAME_CHARS.sub("_", value).strip()
    return sanitized or "guild"


def read_card_guilds(path: Path = CONFIG_PATH) -> list[str]:
    if not path.exists():
        raise CardGenerationError(f"ギルド指定ファイルが見つかりません: {path}")

    guilds = [line.strip() for line in path.read_text(encoding="utf-8-sig").splitlines()]
    guilds = [guild for guild in guilds if guild and not guild.startswith("#")]
    if not guilds:
        raise CardGenerationError(f"ギルド指定ファイルが空です: {path}")
    return guilds


def latest_file(directory: Path, pattern: str) -> Path:
    files = [path for path in directory.glob(pattern) if path.is_file()]
    if not files:
        raise CardGenerationError(f"対象ファイルが見つかりません: {directory / pattern}")
    return max(files, key=lambda path: (path.stat().st_mtime, path.name))


def output_date_from_summary(summary_path: Path) -> str:
    match = SUMMARY_DATE_PATTERN.search(summary_path.name)
    if match:
        return match.group(1)
    return datetime.fromtimestamp(summary_path.stat().st_mtime).strftime("%Y-%m-%d")


def normalize_header(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def worksheet_to_dicts(ws) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_header(cell) for cell in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() != "" for cell in row):
            continue
        record = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        records.append(record)
    return records


def load_summary(summary_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(summary_path, data_only=True, read_only=True)
    try:
        missing = [sheet for sheet in ("guild_metrics", "rankings") if sheet not in wb.sheetnames]
        if missing:
            raise CardGenerationError(f"summary に必要なシートがありません: {', '.join(missing)}")
        return worksheet_to_dicts(wb["guild_metrics"]), worksheet_to_dicts(wb["rankings"])
    finally:
        wb.close()


def row_guild_name(record: dict[str, Any]) -> str:
    for key in ("guild_name", "guild", "name", "ギルド名"):
        value = record.get(key)
        if value is not None and str(value).strip():
            return str(value).strip()
    return ""


def find_guild_record(records: Iterable[dict[str, Any]], guild_name: str) -> dict[str, Any] | None:
    target = guild_name.strip()
    for record in records:
        if row_guild_name(record) == target:
            return dict(record)
    return None


def find_rank_by_avg_cpm(rankings: Iterable[dict[str, Any]], guild_name: str) -> Any:
    record = find_guild_record(rankings, guild_name)
    if not record:
        return None
    for key in ("rank_by_avg_cpm", "avg_cpm_rank", "rank", "順位"):
        if key in record:
            return record[key]
    return None


def find_latest_guild_workbook(guild_name: str) -> Path | None:
    guild_dir = DATA_DIR / guild_name
    if not guild_dir.exists():
        logger.warning("data フォルダが見つからないためスキップします: %s", guild_dir)
        return None
    try:
        return latest_file(guild_dir, GUILD_PATTERN)
    except CardGenerationError as exc:
        logger.warning("%s", exc)
        return None


def load_members(guild_workbook_path: Path) -> list[dict[str, Any]] | None:
    wb = load_workbook(guild_workbook_path, data_only=True, read_only=True)
    try:
        if "members" not in wb.sheetnames:
            logger.warning("members シートがないためスキップします: %s", guild_workbook_path)
            return None
        members = worksheet_to_dicts(wb["members"])
    finally:
        wb.close()

    members.sort(key=lambda record: numeric_sort_key(record.get("rank")))
    return members


def numeric_sort_key(value: Any) -> tuple[int, str]:
    number = to_number(value)
    if number is None:
        return (10**9, str(value or ""))
    return (int(number), str(value or ""))


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return False


def to_number(value: Any) -> float | None:
    if is_blank(value):
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    text = str(value).strip().replace(",", "").replace("%", "")
    try:
        return float(text)
    except ValueError:
        return None


def format_integer(value: Any, comma: bool = False) -> str:
    number = to_number(value)
    if number is None:
        return ""
    formatted = f"{number:,.0f}" if comma else f"{number:.0f}"
    return formatted


def format_rate(value: Any) -> str:
    number = to_number(value)
    if number is None:
        return ""
    if abs(number) > 1:
        number = number / 100
    return f"{number * 100:.2f}%"


def format_date_or_text(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    return str(value)


def format_date_only(value: Any) -> str:
    if is_blank(value):
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    text = str(value).strip()
    match = re.match(r"(\d{4}-\d{2}-\d{2})", text)
    if match:
        return match.group(1)
    return text


def format_placeholder_value(key: str, raw_value: Any) -> str:
    if key == "retrieved_at":
        return format_date_only(raw_value)
    if is_blank(raw_value):
        return ""
    if key in CPM_KEYS:
        return format_integer(raw_value, comma=True)
    if key in COUNT_KEYS:
        return format_integer(raw_value, comma=False)
    if key in RATE_KEYS:
        return format_rate(raw_value)
    return format_date_or_text(raw_value)


def build_context(guild_name: str, metric_record: dict[str, Any], rank_by_avg_cpm: Any) -> dict[str, str]:
    raw_context = {str(key): value for key, value in metric_record.items()}
    raw_context["guild_name"] = guild_name
    raw_context["rank_by_avg_cpm"] = rank_by_avg_cpm

    expected_keys = set(raw_context) | CPM_KEYS | COUNT_KEYS | RATE_KEYS | {
        "retrieved_at",
        "currently_holding",
        "auto_comment",
        "autocomment",
        "ai_comment",
    }
    return {key: format_placeholder_value(key, raw_context.get(key)) for key in expected_keys}


def replace_placeholders(wb, context: dict[str, str]) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str) or "{{" not in cell.value:
                    continue

                exact_match = EXACT_PLACEHOLDER_PATTERN.match(cell.value)
                if exact_match:
                    key = exact_match.group(1)
                    value = context.get(key, "")
                    if key in NUMERIC_PLACEHOLDER_KEYS and value != "":
                        number = to_number(value)
                        if number is not None:
                            cell.value = int(number)
                            cell.number_format = "0"
                            continue
                    cell.value = value
                    continue

                def replacement(match: re.Match[str]) -> str:
                    key = match.group(1)
                    return context.get(key, "")

                cell.value = PLACEHOLDER_PATTERN.sub(replacement, cell.value)


def write_autocomment(ws, context: dict[str, str]) -> None:
    """Write the generated auto comment into the fixed card comment cell."""
    ws["B26"] = (
        context.get("auto_comment")
        or context.get("autocomment")
        or context.get("ai_comment")
        or ""
    )


def find_member_blocks(ws) -> tuple[tuple[int, int, int, int], tuple[int, int, int, int]]:
    """Find left/right member table header blocks in the existing template.

    Returns two tuples: (start_row, no_col, name_col, cpm_col). The template is
    expected to have two `No / 家門名 / CPM` header sets; if it cannot be found,
    the historical B:D and F:H layout is used as a safe fallback.
    """
    found_blocks: list[tuple[int, int, int, int]] = []
    max_header_scan_row = min(ws.max_row, 30)

    for row_index in range(1, max_header_scan_row + 1):
        for col_index in range(1, max(ws.max_column - 1, 1)):
            no_value = normalize_header(ws.cell(row=row_index, column=col_index).value)
            name_value = normalize_header(ws.cell(row=row_index, column=col_index + 1).value)
            cpm_value = normalize_header(ws.cell(row=row_index, column=col_index + 2).value).upper()
            if (no_value, name_value, cpm_value) == MEMBER_HEADER_LABELS:
                found_blocks.append((row_index + 1, col_index, col_index + 1, col_index + 2))

        if len(found_blocks) >= 2:
            break

    if len(found_blocks) >= 2:
        return found_blocks[0], found_blocks[1]

    logger.warning(
        "個人戦闘力一覧シートの見出しを2セット検出できなかったため、B:D / F:H を使用します。"
    )
    return (2, *MEMBER_LEFT_COLUMNS), (2, *MEMBER_RIGHT_COLUMNS)


def clear_member_values(ws, blocks: tuple[tuple[int, int, int, int], tuple[int, int, int, int]]) -> None:
    for start_row, no_col, name_col, cpm_col in blocks:
        for offset in range(MEMBER_MAX_ROWS_PER_SIDE):
            row_index = start_row + offset
            for col_index in (no_col, name_col, cpm_col):
                ws.cell(row=row_index, column=col_index).value = None


def write_member_row(ws, row_index: int, columns: tuple[int, int, int], member: dict[str, Any], fallback_rank: int) -> None:
    no_col, name_col, cpm_col = columns
    rank_value = to_number(member.get("rank"))
    player_name = format_date_or_text(member.get("player_name")).strip()

    ws.cell(row=row_index, column=no_col, value=int(rank_value) if rank_value is not None else fallback_rank)
    ws.cell(row=row_index, column=name_col, value=player_name)
    cpm_cell = ws.cell(row=row_index, column=cpm_col, value=to_number(member.get("cpm")))
    cpm_cell.number_format = "#,##0"


def write_members(ws, members: list[dict[str, Any]]) -> None:
    """Write member names and CPM only, preserving the template No columns."""
    member_rows = members[:MEMBER_MAX_COUNT]

    # The template already has No values in B3:B27 and E3:E27. Do not clear or
    # write those columns; only replace the data cells so formatting/layout stays
    # exactly as the template defines it.
    for row_index in range(3, 28):
        for col_index in (3, 4, 6, 7):
            ws.cell(row=row_index, column=col_index).value = None

    for index, member in enumerate(member_rows):
        if index < MEMBER_MAX_ROWS_PER_SIDE:
            row_index = 3 + index
            name_col = 3
            cpm_col = 4
        else:
            row_index = 3 + (index - MEMBER_MAX_ROWS_PER_SIDE)
            name_col = 6
            cpm_col = 7

        ws.cell(row=row_index, column=name_col, value=format_date_or_text(member.get("player_name")).strip())
        cpm_cell = ws.cell(row=row_index, column=cpm_col, value=to_number(member.get("cpm")))
        cpm_cell.number_format = "#,##0"


def int_for_chart(value: Any) -> int:
    number = to_number(value)
    if number is None:
        return 0
    return int(round(number))


def add_cpm_band_chart(wb, context: dict[str, str]) -> None:
    """Draw the CPM band chart with cell fills instead of an Excel chart object."""
    card_ws = wb[TEMPLATE_CARD_SHEET]
    card_ws._charts = []

    counts = [int_for_chart(context.get(key)) for _, key in CHART_BANDS]
    max_count = max(counts) if counts else 0
    bar_fill = PatternFill("solid", fgColor=PINK_FILL)
    empty_fill = PatternFill("solid", fgColor=LIGHT_PINK_FILL)
    thin_side = Side(style="hair", color=GRAPH_BORDER)
    subtle_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for index, ((label, _), count) in enumerate(zip(CHART_BANDS, counts)):
        row_index = GRAPH_START_ROW + index
        card_ws.cell(row=row_index, column=GRAPH_LABEL_COL, value=label)
        count_cell = card_ws.cell(row=row_index, column=GRAPH_COUNT_COL, value=count)
        count_cell.number_format = "#,##0"

        filled_cells = 0
        if max_count > 0 and count > 0:
            filled_cells = max(1, math.ceil(count / max_count * GRAPH_WIDTH))

        for offset, col_index in enumerate(range(GRAPH_START_COL, GRAPH_END_COL + 1), start=1):
            cell = card_ws.cell(row=row_index, column=col_index)
            cell.value = None
            cell.fill = bar_fill if offset <= filled_cells else empty_fill
            cell.border = subtle_border


def configure_excel_png_view(excel, worksheet) -> None:
    """Switch Excel away from page-break preview before copying a range."""
    worksheet.Activate()
    try:  # pragma: no cover - depends on Excel COM.
        excel.ActiveWindow.View = -4143  # xlNormalView
        excel.ActiveWindow.DisplayGridlines = False
        excel.ActiveWindow.DisplayHeadings = False
    except Exception as exc:
        logger.warning("Excel表示設定の変更に失敗しました: %s", exc)
    try:  # pragma: no cover - depends on Excel COM.
        worksheet.DisplayPageBreaks = False
    except Exception as exc:
        logger.warning("改ページ表示の無効化に失敗しました: %s", exc)


def export_sheet_range_to_png(excel, worksheet, cell_range: str, png_path: Path) -> None:
    """Export one worksheet range as PNG through an ephemeral Excel chart object."""
    configure_excel_png_view(excel, worksheet)
    source_range = worksheet.Range(cell_range)
    source_range.CopyPicture(Appearance=1, Format=2)

    chart_object = worksheet.ChartObjects().Add(
        source_range.Left,
        source_range.Top,
        source_range.Width,
        source_range.Height,
    )
    try:
        chart_object.Activate()
        chart_object.Chart.Paste()
        chart_object.Chart.Export(str(png_path.resolve()))
    finally:
        chart_object.Delete()


def export_png_with_excel(xlsx_path: Path, karte_png_path: Path, members_png_path: Path) -> None:
    """Best-effort Windows Excel COM export of the card and member sheets to PNG."""
    if platform.system() != "Windows":
        logger.warning("PNG出力はWindows + Excel環境のみ対応のためスキップします: %s", xlsx_path)
        return

    if importlib.util.find_spec("pythoncom") is None or importlib.util.find_spec("win32com.client") is None:
        logger.warning("pywin32 を利用できないため PNG 出力をスキップします: %s", xlsx_path)
        return

    pythoncom = importlib.import_module("pythoncom")
    win32com_client = importlib.import_module("win32com.client")

    excel = None
    workbook = None
    try:  # pragma: no cover - depends on Excel COM.
        pythoncom.CoInitialize()
        excel = win32com_client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(str(xlsx_path.resolve()))

        exports = (
            (TEMPLATE_CARD_SHEET, "A1:M30", karte_png_path),
            (TEMPLATE_MEMBERS_SHEET, "A1:G28", members_png_path),
        )
        for sheet_name, cell_range, png_path in exports:
            try:
                sheet = workbook.Worksheets(sheet_name)
                export_sheet_range_to_png(excel, sheet, cell_range, png_path)
                logger.info("PNG を出力しました: %s", png_path)
            except Exception as exc:
                logger.warning("PNG 出力に失敗しました（xlsx は作成済みです）: %s: %s", png_path, exc)
    except Exception as exc:
        logger.warning("PNG 出力に失敗しました（xlsx は作成済みです）: %s", exc)
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        try:
            pythoncom.CoUninitialize()  # type: ignore[name-defined]
        except Exception:
            pass


def create_card_for_guild(
    guild_name: str,
    metric_record: dict[str, Any],
    rank_by_avg_cpm: Any,
    summary_date: str,
) -> Path | None:
    guild_workbook_path = find_latest_guild_workbook(guild_name)
    if guild_workbook_path is None:
        return None

    members = load_members(guild_workbook_path)
    if members is None:
        return None

    if not TEMPLATE_PATH.exists():
        raise CardGenerationError(f"テンプレートが見つかりません: {TEMPLATE_PATH}")

    safe_guild_name = sanitize_filename(guild_name)
    out_dir = OUTPUT_DIR / safe_guild_name
    out_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = out_dir / f"karte_{safe_guild_name}_{summary_date}.xlsx"
    karte_png_path = out_dir / f"karte_{safe_guild_name}_{summary_date}.png"
    members_png_path = out_dir / f"members_{safe_guild_name}_{summary_date}.png"

    shutil.copy2(TEMPLATE_PATH, xlsx_path)
    wb = load_workbook(xlsx_path)
    if TEMPLATE_CARD_SHEET not in wb.sheetnames or TEMPLATE_MEMBERS_SHEET not in wb.sheetnames:
        wb.close()
        raise CardGenerationError(
            f"テンプレートに必要なシートがありません: {TEMPLATE_CARD_SHEET}, {TEMPLATE_MEMBERS_SHEET}"
        )

    try:
        context = build_context(guild_name, metric_record, rank_by_avg_cpm)
        replace_placeholders(wb, context)
        write_autocomment(wb[TEMPLATE_CARD_SHEET], context)
        write_members(wb[TEMPLATE_MEMBERS_SHEET], members)
        add_cpm_band_chart(wb, context)
        wb.save(xlsx_path)
    finally:
        wb.close()

    logger.info("XLSX を出力しました: %s", xlsx_path)
    export_png_with_excel(xlsx_path, karte_png_path, members_png_path)
    return xlsx_path


def main() -> int:
    try:
        ensure_dirs()
        guilds = read_card_guilds()
        summary_path = latest_file(ANALYSIS_DIR, SUMMARY_PATTERN)
        summary_date = output_date_from_summary(summary_path)
        logger.info("summary を読み込みます: %s", summary_path)
        guild_metrics, rankings = load_summary(summary_path)

        created = 0
        for guild_name in guilds:
            metric_record = find_guild_record(guild_metrics, guild_name)
            if metric_record is None:
                logger.warning("summary にギルドが存在しないためスキップします: %s", guild_name)
                continue

            rank_by_avg_cpm = find_rank_by_avg_cpm(rankings, guild_name)
            if rank_by_avg_cpm is None:
                rank_by_avg_cpm = metric_record.get("rank_by_avg_cpm")

            if create_card_for_guild(guild_name, metric_record, rank_by_avg_cpm, summary_date):
                created += 1

        if created == 0:
            logger.warning("作成されたカルテはありません。")
        else:
            logger.info("カルテ作成が完了しました: %d 件", created)
        return 0
    except CardGenerationError as exc:
        logger.error("%s", exc)
        return 1


if __name__ == "__main__":
    sys.exit(main())
