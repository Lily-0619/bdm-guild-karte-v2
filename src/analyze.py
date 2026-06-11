"""Analyze saved DBonk guild Excel exports.

This script is intentionally read-only with respect to DBonk: it does not fetch
web data, does not use Playwright, and only reads already-saved Excel files from
``data/``. Results are written as an Excel workbook under ``analysis/``.
"""

from __future__ import annotations

import json
import logging
import re
import statistics
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from .paths import ANALYSIS_DIR, CONFIG_DIR, DATA_DIR, PROJECT_ROOT, ensure_dirs
except ImportError:  # 直接実行された場合のため
    from paths import ANALYSIS_DIR, CONFIG_DIR, DATA_DIR, PROJECT_ROOT, ensure_dirs  # type: ignore

BASE_DIR = PROJECT_ROOT
SETTINGS_FILE = CONFIG_DIR / "analysis_settings.json"

TOP_AVG_COUNTS = [10, 15, 20, 25]
CPM_BANDS = [
    ("cpm_130000_plus_count", 130000, None),
    ("cpm_120000_129999_count", 120000, 129999),
    ("cpm_110000_119999_count", 110000, 119999),
    ("cpm_100000_109999_count", 100000, 109999),
    ("cpm_90000_99999_count", 90000, 99999),
    ("cpm_under_90000_count", None, 89999),
]
CPM_BANDS_LABEL = (
    "130000+, 120000-129999, 110000-119999, "
    "100000-109999, 90000-99999, under90000"
)

SUMMARY_FIELDS = [
    "total_cp",
    "total_family_cp",
    "active_member_count",
    "low_member_cp",
    "high_member_cp",
    "declared_on_other_guild",
    "declared_by_other_guild",
    "total_war",
    "all_time_win_rate",
    "total_node_wars",
    "node_won",
    "total_siege_wars",
    "siege_won",
    "currently_holding",
]

GUILD_METRICS_COLUMNS = [
    "guild_name",
    "source_file",
    "retrieved_at",
    "member_count",
    "avg_cpm",
    "median_cpm",
    "top10_avg_cpm",
    "top15_avg_cpm",
    "top20_avg_cpm",
    "top25_avg_cpm",
    "cpm_130000_plus_count",
    "cpm_120000_129999_count",
    "cpm_110000_119999_count",
    "cpm_100000_109999_count",
    "cpm_90000_99999_count",
    "cpm_under_90000_count",
    "max_cpm",
    "min_cpm",
    "stdev_cpm",
    "total_cp",
    "total_family_cp",
    "active_member_count",
    "low_member_cp",
    "high_member_cp",
    "declared_on_other_guild",
    "declared_by_other_guild",
    "total_war",
    "all_time_win_rate",
    "total_node_wars",
    "node_won",
    "node_win_rate",
    "total_siege_wars",
    "siege_won",
    "siege_win_rate",
    "node_siege_total",
    "node_siege_win_total",
    "currently_holding",
    "prev_avg_cpm",
    "avg_cpm_growth",
    "avg_cpm_growth_rate",
    "rank_percentile",
    "avg_cpm_diff_from_all_avg",
    "growth_rank",
    "growth_vs_all_avg",
    "power_type",
    "auto_comment",
]

RANKINGS_COLUMNS = [
    "rank_by_avg_cpm",
    "guild_name",
    "avg_cpm",
    "median_cpm",
    "top10_avg_cpm",
    "top15_avg_cpm",
    "top20_avg_cpm",
    "top25_avg_cpm",
    "member_count",
    "stdev_cpm",
    "cpm_130000_plus_count",
    "cpm_120000_129999_count",
    "cpm_110000_119999_count",
    "cpm_100000_109999_count",
    "cpm_90000_99999_count",
    "cpm_under_90000_count",
    "node_win_rate",
    "siege_win_rate",
    "avg_cpm_growth",
    "avg_cpm_growth_rate",
    "rank_percentile",
    "avg_cpm_diff_from_all_avg",
    "growth_rank",
    "growth_vs_all_avg",
    "power_type",
    "auto_comment",
]

DATE_IN_FILENAME_RE = re.compile(r"(\d{4}-\d{2}-\d{2})")

CPM_NUMBER_FORMAT = "#,##0"
COUNT_NUMBER_FORMAT = "0"
PERCENT_NUMBER_FORMAT = "0.00%"

CPM_FORMAT_COLUMNS = {
    "avg_cpm",
    "median_cpm",
    "top10_avg_cpm",
    "top15_avg_cpm",
    "top20_avg_cpm",
    "top25_avg_cpm",
    "max_cpm",
    "min_cpm",
    "stdev_cpm",
    "total_cp",
    "total_family_cp",
    "low_member_cp",
    "high_member_cp",
    "prev_avg_cpm",
    "avg_cpm_growth",
    "avg_cpm_diff_from_all_avg",
    "growth_vs_all_avg",
}

COUNT_FORMAT_COLUMNS = {
    "rank_by_avg_cpm",
    "growth_rank",
    "member_count",
    "active_member_count",
    "declared_on_other_guild",
    "declared_by_other_guild",
    "total_war",
    "total_node_wars",
    "node_won",
    "total_siege_wars",
    "siege_won",
    "node_siege_total",
    "node_siege_win_total",
}
COUNT_FORMAT_COLUMNS.update(band_name for band_name, _, _ in CPM_BANDS)

PERCENT_FORMAT_COLUMNS = {
    "all_time_win_rate",
    "node_win_rate",
    "siege_win_rate",
    "avg_cpm_growth_rate",
    "rank_percentile",
}


@dataclass(frozen=True)
class AnalysisSettings:
    """Configurable analysis parameters."""

    top_avg_counts: list[int]


def setup_logging() -> None:
    """Configure console logging for PowerShell, macOS Terminal, and CI."""

    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")


def load_settings() -> AnalysisSettings:
    """Load optional future-facing settings from config/analysis_settings.json.

    The script works without the JSON file. If present, it can contain:
    {
      "top_avg_counts": [10, 15, 20, 25]
    }
    """

    settings = AnalysisSettings(top_avg_counts=TOP_AVG_COUNTS.copy())
    if not SETTINGS_FILE.exists():
        return settings

    try:
        raw = json.loads(SETTINGS_FILE.read_text(encoding="utf-8"))
        top_counts = raw.get("top_avg_counts", settings.top_avg_counts)
        return AnalysisSettings(top_avg_counts=[int(value) for value in top_counts])
    except Exception as exc:  # noqa: BLE001 - bad config should not stop analysis.
        logging.warning("設定ファイルの読み込みに失敗しました。定数を使います: %s", exc)
        return settings


def to_number(value: Any) -> float | int | None:
    """Convert Excel values such as '120,000' or '52.3%' to numbers."""

    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int | float):
        return value

    text = str(value).strip()
    if not text:
        return None

    is_percent = text.endswith("%")
    text = text.rstrip("%").replace(",", "")
    try:
        number = float(text)
    except ValueError:
        return None

    if is_percent:
        return number / 100
    if number.is_integer():
        return int(number)
    return number


def numeric_or_blank(value: Any) -> float | int | str:
    """Return a numeric value when possible, otherwise keep strings as-is."""

    number = to_number(value)
    if number is not None:
        return number
    if value is None:
        return ""
    return value


def round_numeric(value: Any) -> int | str:
    """Round a numeric value to an integer while preserving blanks."""

    number = to_number(value)
    if number is None:
        return "" if value is None else value
    return round(number)


def normalize_rate(value: Any) -> float | str:
    """Return a ratio suitable for Excel percentage formatting.

    Values already stored as ratios (for example 0.5238) are kept as-is. Values
    that appear to be percentages (for example 52.38) are converted to 0.5238.
    """

    number = to_number(value)
    if number is None:
        return ""
    if abs(number) > 1:
        return number / 100
    return number


def sheet_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Read an Excel sheet into dictionaries keyed by the header row."""

    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"必要なシートが見つかりません: {sheet_name}")

        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        headers = next(rows, None)
        if not headers:
            return []

        keys = [str(header).strip() if header is not None else "" for header in headers]
        return [
            {key: value for key, value in zip(keys, row, strict=False) if key}
            for row in rows
            if any(cell is not None for cell in row)
        ]
    finally:
        workbook.close()


def read_cpm_values(path: Path) -> list[float]:
    """Read numeric CPM values from the members sheet."""

    cpms: list[float] = []
    for row in sheet_rows(path, "members"):
        cpm = to_number(row.get("cpm"))
        if cpm is not None:
            cpms.append(float(cpm))
    return cpms


def average(values: Iterable[float]) -> float | str:
    """Return the arithmetic mean, or blank for an empty sequence."""

    values_list = list(values)
    if not values_list:
        return ""
    return sum(values_list) / len(values_list)


def rounded_average(values: Iterable[float]) -> int | str:
    """Return a rounded arithmetic mean for CPM-oriented metrics."""

    value = average(values)
    if value == "":
        return ""
    return round(value)


def safe_rate(numerator: Any, denominator: Any) -> float | str:
    """Divide while avoiding zero division and blank/non-numeric denominators."""

    num = to_number(numerator) or 0
    den = to_number(denominator) or 0
    if den == 0:
        return ""
    return num / den


def workbook_sort_key(path: Path) -> tuple[datetime, float, str]:
    """Sort workbooks by date in filename first, then mtime and filename."""

    match = DATE_IN_FILENAME_RE.search(path.name)
    if match:
        file_date = datetime.strptime(match.group(1), "%Y-%m-%d")
    else:
        file_date = datetime.min
    return (file_date, path.stat().st_mtime, path.name)


def find_guild_workbooks(guild_dir: Path) -> list[Path]:
    """Return guild_*.xlsx workbooks in chronological order."""

    files = [
        path
        for path in guild_dir.glob("guild_*.xlsx")
        if path.is_file() and not path.name.startswith("~$")
    ]
    return sorted(files, key=workbook_sort_key)


def calculate_cpm_metrics(cpms: list[float], settings: AnalysisSettings) -> dict[str, Any]:
    """Calculate CPM statistics for one guild."""

    metrics: dict[str, Any] = {"member_count": len(cpms)}
    if not cpms:
        metrics.update(
            {
                "avg_cpm": "",
                "median_cpm": "",
                "max_cpm": "",
                "min_cpm": "",
                "stdev_cpm": 0,
            }
        )
    else:
        metrics.update(
            {
                "avg_cpm": round(statistics.mean(cpms)),
                "median_cpm": round(statistics.median(cpms)),
                "max_cpm": round(max(cpms)),
                "min_cpm": round(min(cpms)),
                "stdev_cpm": round(statistics.pstdev(cpms)) if len(cpms) > 1 else 0,
            }
        )

    sorted_cpms = sorted(cpms, reverse=True)
    for count in settings.top_avg_counts:
        metrics[f"top{count}_avg_cpm"] = rounded_average(sorted_cpms[:count])
    metrics.update(calculate_cpm_band_counts(cpms))

    return metrics


def calculate_cpm_band_counts(cpms: list[float]) -> dict[str, int]:
    """Count members in non-overlapping CPM bands."""

    band_counts: dict[str, int] = {}
    for band_name, lower_bound, upper_bound in CPM_BANDS:
        count = 0
        for cpm in cpms:
            if lower_bound is not None and cpm < lower_bound:
                continue
            if upper_bound is not None and cpm > upper_bound:
                continue
            count += 1
        band_counts[band_name] = count
    return band_counts


def read_summary_values(path: Path) -> dict[str, Any]:
    """Read the first row of the summary sheet."""

    rows = sheet_rows(path, "summary")
    if not rows:
        return {}
    return rows[0]


def calculate_previous_average(previous_file: Path | None) -> float | str:
    """Calculate previous average CPM if a previous workbook exists."""

    if previous_file is None:
        return ""
    previous_cpms = read_cpm_values(previous_file)
    previous_average = rounded_average(previous_cpms)
    return previous_average


def analyze_guild(guild_dir: Path, settings: AnalysisSettings) -> dict[str, Any]:
    """Analyze the latest workbook in one guild directory."""

    workbooks = find_guild_workbooks(guild_dir)
    if not workbooks:
        raise FileNotFoundError(f"guild_*.xlsx が見つかりません: {guild_dir}")

    latest_file = workbooks[-1]
    previous_file = workbooks[-2] if len(workbooks) >= 2 else None

    cpms = read_cpm_values(latest_file)
    summary = read_summary_values(latest_file)
    cpm_metrics = calculate_cpm_metrics(cpms, settings)

    guild_name = summary.get("guild_name") or guild_dir.name
    retrieved_at = summary.get("retrieved_at")
    if not retrieved_at:
        member_rows = sheet_rows(latest_file, "members")
        retrieved_at = member_rows[0].get("retrieved_at", "") if member_rows else ""

    metrics: dict[str, Any] = {
        "guild_name": guild_name,
        "source_file": str(latest_file.relative_to(BASE_DIR)),
        "retrieved_at": retrieved_at,
        **cpm_metrics,
    }

    for field in SUMMARY_FIELDS:
        metrics[field] = numeric_or_blank(summary.get(field))

    for field in CPM_FORMAT_COLUMNS | COUNT_FORMAT_COLUMNS:
        if field in metrics:
            metrics[field] = round_numeric(metrics[field])

    metrics["all_time_win_rate"] = normalize_rate(metrics.get("all_time_win_rate"))

    metrics["node_win_rate"] = safe_rate(
        metrics.get("node_won"), metrics.get("total_node_wars")
    )
    metrics["siege_win_rate"] = safe_rate(
        metrics.get("siege_won"), metrics.get("total_siege_wars")
    )

    total_node_wars = to_number(metrics.get("total_node_wars")) or 0
    total_siege_wars = to_number(metrics.get("total_siege_wars")) or 0
    node_won = to_number(metrics.get("node_won")) or 0
    siege_won = to_number(metrics.get("siege_won")) or 0
    metrics["node_siege_total"] = total_node_wars + total_siege_wars
    metrics["node_siege_win_total"] = node_won + siege_won

    prev_avg_cpm = calculate_previous_average(previous_file)
    metrics["prev_avg_cpm"] = prev_avg_cpm
    if prev_avg_cpm != "" and metrics["avg_cpm"] != "":
        metrics["avg_cpm_growth"] = round_numeric(metrics["avg_cpm"] - prev_avg_cpm)
        metrics["avg_cpm_growth_rate"] = safe_rate(
            metrics["avg_cpm_growth"], prev_avg_cpm
        )
    else:
        metrics["avg_cpm_growth"] = ""
        metrics["avg_cpm_growth_rate"] = ""

    return metrics


def collect_metrics(settings: AnalysisSettings) -> tuple[list[dict[str, Any]], list[tuple[str, str]]]:
    """Analyze every guild directory under data/."""

    if not DATA_DIR.exists():
        logging.warning("data フォルダが見つかりません: %s", DATA_DIR)
        return [], []

    metrics: list[dict[str, Any]] = []
    failures: list[tuple[str, str]] = []
    guild_dirs = sorted([path for path in DATA_DIR.iterdir() if path.is_dir()], key=lambda p: p.name)

    for guild_dir in guild_dirs:
        try:
            guild_metrics = analyze_guild(guild_dir, settings)
            metrics.append(guild_metrics)
            logging.info("読み込み成功: %s", guild_metrics["guild_name"])
        except Exception as exc:  # noqa: BLE001 - continue with the next guild.
            failures.append((guild_dir.name, str(exc)))
            logging.exception("読み込み失敗: %s", guild_dir.name)

    return metrics, failures


def append_table(worksheet: Any, columns: list[str], rows: list[dict[str, Any]]) -> None:
    """Append a header row and dictionary rows to a worksheet."""

    worksheet.append(columns)
    for row in rows:
        worksheet.append([row.get(column, "") for column in columns])
    style_header(worksheet)
    apply_number_formats(worksheet, columns)


def style_header(worksheet: Any) -> None:
    """Apply light formatting that remains compatible with Excel on Windows/Mac."""

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_columns(worksheet)


def autosize_columns(worksheet: Any) -> None:
    """Set practical column widths for readability."""

    for column_cells in worksheet.columns:
        column_letter = get_column_letter(column_cells[0].column)
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 45)


def apply_number_formats(worksheet: Any, columns: list[str]) -> None:
    """Apply Excel number formats by column while keeping values numeric."""

    for column_index, column_name in enumerate(columns, start=1):
        if column_name in CPM_FORMAT_COLUMNS:
            number_format = CPM_NUMBER_FORMAT
        elif column_name in COUNT_FORMAT_COLUMNS:
            number_format = COUNT_NUMBER_FORMAT
        elif column_name in PERCENT_FORMAT_COLUMNS:
            number_format = PERCENT_NUMBER_FORMAT
        else:
            continue

        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            if cell.value == "" or cell.value is None:
                continue
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal="right")


def median_numeric(values: Iterable[Any]) -> float:
    """Return the median of numeric values, or 0 for an empty sequence."""

    numbers = [float(number) for value in values if (number := to_number(value)) is not None]
    if not numbers:
        return 0
    return statistics.median(numbers)


def rank_position_label(rank_percentile: Any) -> str:
    """Convert rank percentile into a readable position label."""

    percentile = to_number(rank_percentile)
    if percentile is None:
        return "未判定"
    if percentile <= 0.20:
        return "上位帯"
    if percentile <= 0.50:
        return "中位上位"
    if percentile <= 0.80:
        return "中位"
    return "下位寄り"


def classify_power_type(row: dict[str, Any], stdev_median: float) -> str:
    """Classify guild power composition using fixed, non-AI rules."""

    avg_cpm = to_number(row.get("avg_cpm")) or 0
    median_cpm = to_number(row.get("median_cpm")) or 0
    top10_avg_cpm = to_number(row.get("top10_avg_cpm")) or avg_cpm
    stdev_cpm = to_number(row.get("stdev_cpm")) or 0
    member_count = to_number(row.get("member_count")) or 0
    top_gap = top10_avg_cpm - avg_cpm
    median_gap = abs(avg_cpm - median_cpm)
    high_count = (to_number(row.get("cpm_130000_plus_count")) or 0) + (
        to_number(row.get("cpm_120000_129999_count")) or 0
    )
    middle_count = (to_number(row.get("cpm_110000_119999_count")) or 0) + (
        to_number(row.get("cpm_100000_109999_count")) or 0
    )
    under_90k_count = to_number(row.get("cpm_under_90000_count")) or 0

    if member_count and under_90k_count / member_count >= 0.30:
        return "発展途上型"
    if stdev_cpm > stdev_median and top_gap >= max(5000, avg_cpm * 0.06):
        return "上位依存型"
    if member_count and high_count / member_count >= 0.25:
        return "エース層強め"
    if member_count and middle_count / member_count >= 0.45:
        return "中間層厚め"
    if stdev_cpm <= stdev_median and median_gap <= max(2500, avg_cpm * 0.03):
        return "戦力均一型"
    return "中間層厚め"


def describe_power_features(row: dict[str, Any], stdev_median: float) -> str:
    """Build a short feature phrase from CPM distribution metrics."""

    power_type = str(row.get("power_type") or "未判定")
    avg_cpm = to_number(row.get("avg_cpm")) or 0
    top10_avg_cpm = to_number(row.get("top10_avg_cpm")) or avg_cpm
    top25_avg_cpm = to_number(row.get("top25_avg_cpm")) or avg_cpm
    stdev_cpm = to_number(row.get("stdev_cpm")) or 0
    top_gap = top10_avg_cpm - avg_cpm
    top25_gap = top25_avg_cpm - avg_cpm

    if power_type == "エース層強め":
        return "120k以上の人数が目立ち、上位10人平均の圧力が強い構成です"
    if power_type == "中間層厚め":
        return "100k〜119k帯が厚く、上位25人平均も土台を支えています"
    if power_type == "戦力均一型":
        return "上位25人平均と平均との差が控えめで、ばらつきは小さい傾向です"
    if power_type == "上位依存型" or top_gap >= max(5000, avg_cpm * 0.06):
        return "上位10人平均が高く、上位層への依存がやや見られます"
    if power_type == "発展途上型":
        return "90k未満の層もあり、上位25人平均との底上げ余地があります"
    if stdev_cpm > stdev_median or top25_gap > max(2500, avg_cpm * 0.03):
        return "標準偏差はやや大きく、戦力差が出やすい構成です"
    return "平均と中央値の差は大きくなく、まとまりのある構成です"


def describe_growth(row: dict[str, Any]) -> str:
    """Build a growth comparison phrase."""

    growth = to_number(row.get("avg_cpm_growth"))
    growth_vs_all_avg = to_number(row.get("growth_vs_all_avg"))
    if growth is None or row.get("avg_cpm_growth") == "":
        return "前回データがないため、成長比較は未判定です。"

    sign = "+" if growth >= 0 else ""
    if growth_vs_all_avg is None:
        comparison = "全体平均との比較は未判定です"
    elif growth_vs_all_avg > 0:
        comparison = "全体平均を上回る成長が見られます"
    elif growth_vs_all_avg < 0:
        comparison = "全体平均をやや下回る伸びです"
    else:
        comparison = "全体平均並みの伸びです"
    return f"前回比では平均CPMが{sign}{growth:,.0f}で、{comparison}。"


def build_auto_comment(row: dict[str, Any], total_guilds: int, stdev_median: float) -> str:
    """Generate a short rule-based, non-AI analysis comment."""

    rank = to_number(row.get("rank_by_avg_cpm"))
    position = rank_position_label(row.get("rank_percentile"))
    if rank is None:
        opening = f"平均CPMは全体{total_guilds}ギルド中で{position}に位置します。"
    else:
        opening = f"平均CPMは全体{total_guilds}ギルド中{int(rank)}位で、{position}に位置します。"
    feature = describe_power_features(row, stdev_median)
    growth = describe_growth(row)
    return f"{opening}{feature}{growth}"


def enrich_comparison_metrics(metrics: list[dict[str, Any]]) -> None:
    """Add rankings, percentile, growth comparison, type, and comments in-place."""

    total_guilds = len(metrics)
    if total_guilds == 0:
        return

    ranked_rows = sorted(
        metrics,
        key=lambda row: to_number(row.get("avg_cpm")) if to_number(row.get("avg_cpm")) is not None else -1,
        reverse=True,
    )
    avg_cpm_all = average(
        float(avg_cpm) for row in metrics if (avg_cpm := to_number(row.get("avg_cpm"))) is not None
    )
    if avg_cpm_all == "":
        avg_cpm_all = 0

    growth_rows = [row for row in metrics if to_number(row.get("avg_cpm_growth")) is not None and row.get("avg_cpm_growth") != ""]
    avg_growth_all = average(float(to_number(row.get("avg_cpm_growth")) or 0) for row in growth_rows)
    stdev_median = median_numeric(row.get("stdev_cpm") for row in metrics)

    for rank, row in enumerate(ranked_rows, start=1):
        row["rank_by_avg_cpm"] = rank
        row["rank_percentile"] = rank / total_guilds
        avg_cpm = to_number(row.get("avg_cpm"))
        row["avg_cpm_diff_from_all_avg"] = round_numeric(avg_cpm - avg_cpm_all) if avg_cpm is not None else ""

    sorted_growth_rows = sorted(
        growth_rows,
        key=lambda row: to_number(row.get("avg_cpm_growth")) or 0,
        reverse=True,
    )
    for growth_rank, row in enumerate(sorted_growth_rows, start=1):
        row["growth_rank"] = growth_rank
        growth = to_number(row.get("avg_cpm_growth"))
        row["growth_vs_all_avg"] = (
            round_numeric(growth - avg_growth_all) if growth is not None and avg_growth_all != "" else ""
        )

    for row in metrics:
        row.setdefault("growth_rank", "")
        row.setdefault("growth_vs_all_avg", "")
        row["power_type"] = classify_power_type(row, stdev_median)
        row["auto_comment"] = build_auto_comment(row, total_guilds, stdev_median)


def build_rankings(metrics: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Build average-CPM descending rankings."""

    sorted_metrics = sorted(
        metrics,
        key=lambda row: to_number(row.get("avg_cpm")) if to_number(row.get("avg_cpm")) is not None else -1,
        reverse=True,
    )
    rankings: list[dict[str, Any]] = []
    for rank, row in enumerate(sorted_metrics, start=1):
        ranking_row = {column: row.get(column, "") for column in RANKINGS_COLUMNS}
        ranking_row["rank_by_avg_cpm"] = row.get("rank_by_avg_cpm", rank)
        rankings.append(ranking_row)
    return rankings


def daily_summary_output_file(now: datetime | None = None) -> Path:
    """Return the one-file-per-day summary path.

    The analyzer intentionally uses only the execution date in the filename.
    Running multiple times on the same day writes the same path and overwrites
    that day's workbook with the latest analysis. Running on a different day
    naturally creates a different dated workbook.
    """

    run_datetime = now or datetime.now()
    return ANALYSIS_DIR / f"summary_{run_datetime.date().isoformat()}.xlsx"


def write_analysis_workbook(
    metrics: list[dict[str, Any]], settings: AnalysisSettings, output_file: Path
) -> None:
    """Write guild metrics, rankings, and analysis settings to Excel.

    ``output_file`` is expected to be the daily summary path. ``Workbook.save``
    writes to that exact path, so same-day reruns replace the existing workbook
    instead of creating timestamped duplicates.
    """

    ANALYSIS_DIR.mkdir(parents=True, exist_ok=True)
    if output_file.exists():
        logging.info("同日分析ファイルを上書きします: %s", output_file)

    workbook = Workbook()
    guild_metrics_sheet = workbook.active
    guild_metrics_sheet.title = "guild_metrics"
    append_table(guild_metrics_sheet, GUILD_METRICS_COLUMNS, metrics)

    rankings_sheet = workbook.create_sheet("rankings")
    append_table(rankings_sheet, RANKINGS_COLUMNS, build_rankings(metrics))

    settings_sheet = workbook.create_sheet("settings")
    settings_rows = [
        {"setting": "generated_at", "value": datetime.now().isoformat(timespec="seconds")},
        {"setting": "source_data_dir", "value": str(DATA_DIR.relative_to(BASE_DIR))},
        {"setting": "cpm_bands", "value": CPM_BANDS_LABEL},
        {"setting": "top_avg_counts", "value": ", ".join(map(str, settings.top_avg_counts))},
    ]
    append_table(settings_sheet, ["setting", "value"], settings_rows)

    workbook.save(output_file)


def main() -> None:
    """Run the guild analysis."""

    ensure_dirs()
    setup_logging()
    settings = load_settings()
    metrics, failures = collect_metrics(settings)
    enrich_comparison_metrics(metrics)

    output_file = daily_summary_output_file()
    write_analysis_workbook(metrics, settings, output_file)

    print(f"読み込み成功ギルド数: {len(metrics)}")
    print(f"読み込み失敗ギルド数: {len(failures)}")
    print(f"出力ファイルパス: {output_file}")


if __name__ == "__main__":
    main()
