"""Import existing guild Excel files into SQLite.

Usage:
    python src/import_excel_to_db.py
    python -m src.import_excel_to_db

By default this scans ``data/**/guild_*.xlsx`` and writes to
``data/bdm_guild.sqlite3``. Paths are relative to the project root, so the
same command works on Windows and macOS.
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from .paths import DATA_DIR, PROJECT_ROOT
except ImportError:  # 直接実行された場合のため
    from paths import DATA_DIR, PROJECT_ROOT  # type: ignore

if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src import db  # noqa: E402

DEFAULT_DATA_DIR = DATA_DIR
MEMBER_SHEET_CANDIDATES = {"members", "member", "guild", "ギルド", "メンバー", "membersheet"}
SUMMARY_SHEET_CANDIDATES = {"summary", "サマリー", "集計", "分析"}
NODE_HISTORY_SHEET_CANDIDATES = {"nodehistory", "node_history", "拠点戦履歴", "履歴"}


def _normalize_header_literal(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[\s_\-()（）#＃]", "", str(value).strip().lower())


RAW_MEMBER_HEADER_ALIASES = {
    "rank_no": {"rank", "rank_no", "no", "順位", "ランク", "順位no", "rank no"},
    "class_name": {"class", "class_name", "職", "職業", "クラス", "職名"},
    "family_name": {
        "family",
        "family_name",
        "player_name",
        "player",
        "name",
        "家名",
        "家門名",
        "名前",
        "メンバー名",
    },
    "level": {"level", "lv", "レベル"},
    "cpm": {"cpm", "戦闘力", "cp", "combat power"},
    "fcp": {"fcp", "生活力", "life", "life power", "家門戦闘力"},
}
MEMBER_HEADER_ALIASES = {
    field: {_normalize_header_literal(alias) for alias in aliases}
    for field, aliases in RAW_MEMBER_HEADER_ALIASES.items()
}
RAW_NODE_HEADER_ALIASES = {
    "war_date": {"war_date", "date", "日付", "開催日", "拠点戦日"},
    "node_name": {"node_name", "content_name", "content", "node", "拠点", "拠点名", "場所"},
    "opponent_guild": {"opponent_guild", "opponent", "対戦相手", "相手", "敵ギルド"},
    "result": {"result", "結果", "勝敗"},
}
NODE_HEADER_ALIASES = {
    field: {_normalize_header_literal(alias) for alias in aliases}
    for field, aliases in RAW_NODE_HEADER_ALIASES.items()
}
SUMMARY_KEY_ALIASES = db.SUMMARY_KEY_ALIASES


def main() -> int:
    parser = argparse.ArgumentParser(description="Import guild Excel files into SQLite.")
    parser.add_argument("--data-dir", type=Path, default=DEFAULT_DATA_DIR)
    parser.add_argument("--db-path", type=Path, default=db.DEFAULT_DB_PATH)
    args = parser.parse_args()

    conn = db.initialize(args.db_path)
    files = sorted(args.data_dir.glob("**/guild_*.xlsx"))
    print(f"DB: {args.db_path}")
    print(f"Excel files found: {len(files)}")

    imported = 0
    skipped = 0
    failed = 0
    try:
        for excel_path in files:
            if db.is_file_imported(conn, excel_path):
                print(f"SKIP imported: {excel_path}")
                skipped += 1
                continue
            try:
                result = import_excel_file(conn, excel_path)
                print(
                    "OK imported: "
                    f"{excel_path} guild={result['guild_name']} "
                    f"retrieved_at={result['retrieved_at']} "
                    f"members={result['members']} "
                    f"node_history={result['node_history']} "
                    f"summary={'yes' if result['summary'] else 'no'}"
                )
                imported += 1
            except Exception as exc:  # noqa: BLE001 - log and continue next file
                print(f"ERROR import failed: {excel_path} ({exc})")
                failed += 1
    finally:
        conn.close()

    print(f"Done. imported={imported} skipped={skipped} failed={failed}")
    return 1 if failed else 0


def import_excel_file(conn: Any, excel_path: Path) -> dict[str, Any]:
    """Import one Excel file and mark it as imported on success."""

    snapshot = read_guild_excel(excel_path)
    with conn:
        member_count = db.save_snapshot(conn, **snapshot)
        db.mark_file_imported(conn, excel_path)
    return {
        "guild_name": snapshot["guild_name"],
        "retrieved_at": snapshot["retrieved_at"],
        "members": member_count,
        "node_history": len(snapshot.get("node_history") or []),
        "summary": bool(snapshot.get("summary")),
    }


def read_guild_excel(excel_path: Path) -> dict[str, Any]:
    """Read a guild Excel workbook into the shape expected by src.db."""

    from openpyxl import load_workbook

    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    try:
        guild_name, retrieved_at = infer_metadata(excel_path)
        member_sheet = find_member_sheet(workbook)
        members = read_members(member_sheet)
        node_history = read_node_history(workbook)
        summary = read_summary(workbook)
    finally:
        workbook.close()

    cpm_values = [m["cpm"] for m in members if m.get("cpm") is not None]
    total_cpm = sum(cpm_values) if cpm_values else 0
    avg_cpm = total_cpm / len(cpm_values) if cpm_values else None
    return {
        "retrieved_at": retrieved_at,
        "guild_name": guild_name,
        "members": members,
        "member_count": len(members),
        "avg_cpm": avg_cpm,
        "total_cpm": total_cpm,
        "node_history": node_history,
        "summary": summary,
    }


def infer_metadata(excel_path: Path) -> tuple[str, str]:
    """Infer guild name and retrieved date from guild_<name>_<date>.xlsx."""

    match = re.match(
        r"^guild_(?P<guild>.+)_(?P<date>\d{8}|\d{4}-\d{2}-\d{2}).xlsx$",
        excel_path.name,
    )
    if match:
        raw_date = match.group("date")
        if "-" in raw_date:
            retrieved_at = raw_date
        else:
            retrieved_at = datetime.strptime(raw_date, "%Y%m%d").date().isoformat()
        return match.group("guild"), retrieved_at

    guild_name = excel_path.parent.name if excel_path.parent.name != "data" else "unknown"
    retrieved_at = datetime.fromtimestamp(excel_path.stat().st_mtime).date().isoformat()
    return guild_name, retrieved_at


def find_member_sheet(workbook: Any) -> Any:
    """Return the worksheet that looks most like the member table."""

    best_sheet = workbook.active
    best_score = -1
    for sheet in workbook.worksheets:
        name_score = 2 if normalize_header(sheet.title) in MEMBER_SHEET_CANDIDATES else 0
        try:
            _, header_map = find_header_row(list(sheet.iter_rows(values_only=True)))
            score = name_score + len(header_map)
        except ValueError:
            score = name_score
        if score > best_score:
            best_sheet = sheet
            best_score = score
    return best_sheet


def read_members(sheet: Any) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    header_index, header_map = find_header_row(rows)
    members = []
    for row in rows[header_index + 1 :]:
        member = row_to_member(row, header_map)
        if member.get("family_name"):
            members.append(member)
    return members


def find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    """Find the member table header row using Japanese/English aliases."""

    best_index = -1
    best_map: dict[str, int] = {}
    for index, row in enumerate(rows[:20]):
        normalized_cells = [normalize_header(cell) for cell in row]
        candidate: dict[str, int] = {}
        for field, aliases in MEMBER_HEADER_ALIASES.items():
            for column_index, cell in enumerate(normalized_cells):
                if cell in aliases:
                    candidate[field] = column_index
                    break
        if "family_name" in candidate and len(candidate) > len(best_map):
            best_index = index
            best_map = candidate
    if best_index < 0:
        raise ValueError("member header row was not found")
    return best_index, best_map


def row_to_member(row: tuple[Any, ...], header_map: dict[str, int]) -> dict[str, Any]:
    return {
        "rank_no": to_int(value_at(row, header_map.get("rank_no"))),
        "class_name": clean_text(value_at(row, header_map.get("class_name"))),
        "family_name": clean_text(value_at(row, header_map.get("family_name"))),
        "level": to_int(value_at(row, header_map.get("level"))),
        "cpm": to_float(value_at(row, header_map.get("cpm"))),
        "fcp": to_float(value_at(row, header_map.get("fcp"))),
        "class_name_raw": None,
        "class_name_normalized": None,
        "class_name_version": None,
    }


def read_node_history(workbook: Any) -> list[dict[str, Any]]:
    """Read node_history sheet rows, preserving unrecognized columns as JSON data."""

    sheet = find_sheet_by_candidates(workbook, NODE_HISTORY_SHEET_CANDIDATES)
    if sheet is None:
        return []
    rows = list(sheet.iter_rows(values_only=True))
    header_index, header_map, raw_headers = find_generic_header_row(rows, NODE_HEADER_ALIASES)
    history = []
    for row_no, row in enumerate(rows[header_index + 1 :], start=1):
        raw = row_to_raw_dict(row, raw_headers)
        if not any(value not in (None, "") for value in raw.values()):
            continue
        history.append(
            {
                "row_no": row_no,
                "war_date": clean_text(value_at(row, header_map.get("war_date"))),
                "node_name": clean_text(value_at(row, header_map.get("node_name"))),
                "opponent_guild": clean_text(
                    value_at(row, header_map.get("opponent_guild"))
                ),
                "result": clean_text(value_at(row, header_map.get("result"))),
                "raw": raw,
            }
        )
    return history


def read_summary(workbook: Any) -> dict[str, Any]:
    """Read a summary sheet as a flexible mapping.

    The current scraper writes summary as one header row plus one data row. Older
    or manual sheets may be key/value rows, so both shapes are supported.
    """

    sheet = find_sheet_by_candidates(workbook, SUMMARY_SHEET_CANDIDATES)
    if sheet is None:
        return {}
    rows = list(sheet.iter_rows(values_only=True))
    summary: dict[str, Any] = {}

    for index, row in enumerate(rows[:-1]):
        headers = [clean_text(cell) for cell in row]
        values = rows[index + 1]
        recognized = [h for h in headers if h and canonical_summary_key(h) != h]
        if len(recognized) >= 2:
            for col, header in enumerate(headers):
                if header and col < len(values):
                    summary[canonical_summary_key(header)] = values[col]
            return summary

    for row in rows:
        non_empty = [cell for cell in row if cell not in (None, "")]
        if len(non_empty) >= 2:
            key = clean_text(non_empty[0])
            if key:
                summary[canonical_summary_key(key)] = non_empty[1]
    return summary


def find_sheet_by_candidates(workbook: Any, candidates: set[str]) -> Any | None:
    for sheet in workbook.worksheets:
        if normalize_header(sheet.title) in candidates:
            return sheet
    return None


def find_generic_header_row(
    rows: list[tuple[Any, ...]], aliases_by_field: dict[str, set[str]]
) -> tuple[int, dict[str, int], list[str]]:
    best_index = -1
    best_map: dict[str, int] = {}
    best_headers: list[str] = []
    for index, row in enumerate(rows[:20]):
        normalized_cells = [normalize_header(cell) for cell in row]
        candidate: dict[str, int] = {}
        for field, aliases in aliases_by_field.items():
            for column_index, cell in enumerate(normalized_cells):
                if cell in aliases:
                    candidate[field] = column_index
                    break
        if len(candidate) > len(best_map):
            best_index = index
            best_map = candidate
            best_headers = [clean_text(cell) or f"column_{i + 1}" for i, cell in enumerate(row)]
    if best_index < 0:
        return 0, {}, []
    return best_index, best_map, best_headers


def row_to_raw_dict(row: tuple[Any, ...], headers: list[str]) -> dict[str, Any]:
    raw: dict[str, Any] = {}
    for index, value in enumerate(row):
        key = headers[index] if index < len(headers) else f"column_{index + 1}"
        raw[key] = value
    return raw


def canonical_summary_key(key: str) -> str:
    normalized_key = normalize_key(key)
    for canonical_key, aliases in SUMMARY_KEY_ALIASES.items():
        if normalized_key in {normalize_key(alias) for alias in aliases}:
            return canonical_key
    return key


def value_at(row: tuple[Any, ...], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def normalize_header(value: Any) -> str:
    return _normalize_header_literal(value)


def normalize_key(value: Any) -> str:
    return "".join(ch for ch in str(value).lower() if ch.isalnum())


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def to_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(str(value).replace(",", "").strip()))
    except (TypeError, ValueError):
        return None


def to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").strip().replace("%", ""))
    except (TypeError, ValueError):
        return None


if __name__ == "__main__":
    raise SystemExit(main())
