"""DBonk guild member scraper (MVP, Playwright版)。"""

from __future__ import annotations

import csv
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, List

if TYPE_CHECKING:
    from playwright.sync_api import Locator, Page

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from src import db  # noqa: E402

CONFIG_PATH = PROJECT_ROOT / "config" / "guilds.txt"
DATA_DIR = PROJECT_ROOT / "data"

DBONK_LOGIN_URL = "https://dbonk.com/bdmbsmv2/index.php"
RESULT_TIMEOUT_SECONDS = 40
GUILD_LOAD_TIMEOUT_SECONDS = 60
POLL_INTERVAL_SECONDS = 0.5
DEBUG_SAVE_FILES = False


@dataclass
class MemberRow:
    rank: str
    player_name: str
    level: str
    cpm: str
    fcp: str
    retrieved_at: str


def sanitize_filename(text: str) -> str:
    text = text.strip()
    text = re.sub(r'[\\/:*?"<>|]', "_", text)
    text = re.sub(r"\s+", " ", text)
    text = text.rstrip(". ")
    if not text:
        return "unknown_guild"
    reserved = {
        "CON",
        "PRN",
        "AUX",
        "NUL",
        "COM1",
        "COM2",
        "COM3",
        "COM4",
        "COM5",
        "COM6",
        "COM7",
        "COM8",
        "COM9",
        "LPT1",
        "LPT2",
        "LPT3",
        "LPT4",
        "LPT5",
        "LPT6",
        "LPT7",
        "LPT8",
        "LPT9",
    }
    if text.upper() in reserved:
        text = f"_{text}"
    return text[:80]


def get_guild_data_dir(guild_name: str) -> Path:
    guild_dir = DATA_DIR / sanitize_filename(guild_name)
    guild_dir.mkdir(parents=True, exist_ok=True)
    return guild_dir


def load_guild_names(config_path: Path) -> List[str]:
    if not config_path.exists():
        raise FileNotFoundError(f"ギルド設定ファイルが見つかりません: {config_path}")
    guilds = [
        line.strip()
        for line in config_path.read_text(encoding="utf-8").splitlines()
        if line.strip()
    ]
    if not guilds:
        raise ValueError("guilds.txt が空です。1行に1ギルド名を書いてください。")
    return guilds


def try_auto_login(page: Page) -> None:
    """通常ログインフォームが見えている時だけ .env の情報で自動ログインする。"""
    username_input = page.locator("input[placeholder='Input Username']")
    password_input = page.locator("input[placeholder='Input Password'], input[type='password']")

    if username_input.count() == 0 or password_input.count() == 0:
        print("ログイン済み、または通常ログインフォームが見つからないため自動ログインをスキップします。")
        return

    username = os.getenv("DBONK_USERNAME", "").strip()
    password = os.getenv("DBONK_PASSWORD", "").strip()

    if not username or not password:
        raise RuntimeError(".env に DBONK_USERNAME / DBONK_PASSWORD を設定してください。")

    print("自動ログインを開始します。")
    username_input.first.fill(username)
    password_input.first.fill(password)

    login_btn = page.get_by_role("button", name=re.compile(r"^\s*Login\s*$", re.I))
    if login_btn.count() > 0 and login_btn.first.is_visible():
        login_btn.first.click(timeout=3000)
    else:
        password_input.first.press("Enter")

    page.wait_for_timeout(1500)
    print("自動ログインを実行しました。")


def try_select_asia_server(page: Page) -> None:
    """サーバー選択画面が表示された場合は Asia を試す（失敗しても継続）。"""
    asia_candidates = [
        page.get_by_text(re.compile(r"^\s*Asia\s*$", re.I)),
        page.get_by_role("button", name=re.compile(r"^\s*Asia\s*$", re.I)),
    ]
    for candidate in asia_candidates:
        try:
            if candidate.count() > 0 and candidate.first.is_visible():
                candidate.first.click(timeout=2000)
                page.wait_for_timeout(1000)
                print("Asiaサーバー選択を実行しました。")
                return
        except Exception:
            continue


def get_guild_ranking_search_input(page: Page) -> Locator:
    strict = page.locator('input[placeholder="Search..."]')
    for i in range(strict.count()):
        item = strict.nth(i)
        try:
            if item.is_visible():
                return item
        except Exception:
            continue

    broad = page.locator('input[placeholder*="Search"]')
    for i in range(broad.count()):
        item = broad.nth(i)
        try:
            ph = (item.get_attribute("placeholder") or "").strip()
            if ph == "Search General Chat":
                continue
            if item.is_visible():
                return item
        except Exception:
            continue
    raise RuntimeError("Guild Ranking 画面の Search... 検索欄が見つかりません。")


def close_general_chat_panel(page: Page) -> None:
    try:
        if page.locator('input[placeholder="Search General Chat"]').count() == 0:
            return
    except Exception:
        return
    candidates = [
        page.locator("button:has-text('>')"),
        page.locator("button:has-text('＞')"),
        page.locator("button:has-text('arrow_forward_ios')"),
        page.locator(":is(div,span,i):has-text('>')"),
        page.locator(":is(div,span,i):has-text('＞')"),
        page.locator(":is(div,span,i):has-text('arrow_forward_ios')"),
        page.locator("[aria-label*='chat' i]"),
    ]
    for c in candidates:
        try:
            for i in range(min(c.count(), 10)):
                el = c.nth(i)
                if not el.is_visible():
                    continue
                box = el.bounding_box()
                if box and box["x"] > 650:
                    el.click(timeout=1500)
                    page.wait_for_timeout(400)
                    return
        except Exception:
            continue


def is_on_guild_ranking_page(page: Page) -> bool:
    try:
        if page.get_by_text(re.compile(r"^\s*Guild Rank\s*$", re.I)).first.is_visible():
            return True
    except Exception:
        pass
    try:
        if get_guild_ranking_search_input(page).is_visible():
            return True
    except Exception:
        pass
    return False


def open_left_menu(page: Page) -> None:
    toggles = [
        page.locator("button:has-text('menu')"),
        page.locator("[aria-label*='menu' i]"),
        page.locator("button").filter(has_text=re.compile(r"^\s*menu\s*$", re.I)),
        page.locator(":is(div,span,i)[class*='menu' i]"),
        page.locator(":is(div,span,i):has-text('menu')"),
    ]
    for t in toggles:
        try:
            if t.count() > 0 and t.first.is_visible():
                t.first.click(timeout=2000)
                page.wait_for_timeout(500)
                return
        except Exception:
            continue
    fallback = page.locator(":is(button,div,span,i)").filter(
        has_text=re.compile(r"^\s*menu\s*$", re.I)
    )
    for i in range(min(fallback.count(), 10)):
        el = fallback.nth(i)
        try:
            if not el.is_visible():
                continue
            box = el.bounding_box()
            if box and box["x"] < 220 and box["y"] < 220:
                el.click(timeout=2000)
                page.wait_for_timeout(500)
                return
        except Exception:
            continue


def open_guild_ranking_page(page: Page) -> None:
    if is_on_guild_ranking_page(page):
        return

    close_general_chat_panel(page)
    open_left_menu(page)
    candidates = [
        page.get_by_text("Guild Ranking", exact=True),
        page.locator(":is(a,button,div,span,li):has-text('Guild Ranking')"),
        page.locator("text=Guild Ranking").locator("xpath=parent::*"),
        page.locator("text=Guild Ranking").locator(
            "xpath=ancestor::*[self::a or self::button or self::div or self::span or self::li][1]"
        ),
    ]
    clicked = False
    for c in candidates:
        try:
            limit = min(c.count(), 10)
            for i in range(limit):
                item = c.nth(i)
                if item.is_visible():
                    item.click(timeout=3000)
                    clicked = True
                    break
            if clicked:
                break
        except Exception:
            continue
    if not clicked:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        page.screenshot(path=str(DATA_DIR / "debug_guild_ranking_menu.png"), full_page=True)
        body_text = page.locator("body").inner_text()
        (DATA_DIR / "debug_guild_ranking_menu_text.txt").write_text(
            body_text, encoding="utf-8"
        )
        print("左メニュー内テキスト候補（先頭100行）:")
        for line in [ln.strip() for ln in body_text.splitlines() if ln.strip()][:100]:
            print(f"  - {line}")
        raise RuntimeError("左メニューの Guild Ranking をクリックできませんでした。")

    start = time.monotonic()
    while time.monotonic() - start < RESULT_TIMEOUT_SECONDS:
        try:
            if is_on_guild_ranking_page(page):
                return
        except Exception:
            pass
        time.sleep(POLL_INTERVAL_SECONDS)
    raise RuntimeError("Guild Ranking 画面の読み込み待機に失敗しました。")


def search_guild_in_ranking(page: Page, guild_name: str) -> None:
    search_input = get_guild_ranking_search_input(page)
    search_input.click()
    search_input.fill("")
    search_input.fill(guild_name)
    page.wait_for_timeout(700)


def get_ranking_row_candidates(page: Page) -> List[tuple[Locator, str]]:
    rows = page.locator("table tbody tr, [role='row']")
    result: List[tuple[Locator, str]] = []
    for i in range(rows.count()):
        row = rows.nth(i)
        try:
            if not row.is_visible():
                continue
            text = re.sub(r"\s+", " ", row.inner_text().strip())
            if text:
                result.append((row, text))
        except Exception:
            continue
    return result


def get_visible_ranking_pages(page: Page) -> List[int]:
    pager = page.locator("a, button, span, li, [role='button'], [role='link']")
    pages = set()
    for i in range(pager.count()):
        item = pager.nth(i)
        try:
            if not item.is_visible():
                continue
            text = item.inner_text().strip()
            if text.isdigit():
                n = int(text)
                if 1 <= n <= 200:
                    pages.add(n)
        except Exception:
            continue
    return sorted(pages)


def click_ranking_page(page: Page, page_no: int) -> bool:
    before_rows = [text for _, text in get_ranking_row_candidates(page)[:8]]
    selectors = [
        page.get_by_role("link", name=str(page_no)),
        page.get_by_role("button", name=str(page_no)),
        page.locator(f"[aria-label*='page {page_no}' i]"),
        page.locator(f"[aria-label*='go to page {page_no}' i]"),
        page.locator(":is(a,button,span,li,div,[role='button'],[role='link'])").filter(
            has_text=re.compile(rf"^\s*{page_no}\s*$")
        ),
    ]
    for s in selectors:
        try:
            limit = min(s.count(), 20)
            for i in range(limit):
                el = s.nth(i)
                if not el.is_visible():
                    continue
                clicked = False
                try:
                    el.click(timeout=2500)
                    clicked = True
                except Exception:
                    try:
                        el.evaluate("e => e.click()")
                        clicked = True
                    except Exception:
                        clicked = False
                if not clicked:
                    continue
                page.wait_for_timeout(1000)
                after_rows = [text for _, text in get_ranking_row_candidates(page)[:8]]
                if after_rows and after_rows != before_rows:
                    return True
                if after_rows:
                    return True
        except Exception:
            continue

    print(f"ページ {page_no} のボタン探索に失敗。数字候補一覧:")
    visible_pages = get_visible_ranking_pages(page)
    if visible_pages:
        for candidate_page in visible_pages:
            print(f"  - {candidate_page}")
        return False

    candidates = page.locator(
        ":is(a,button,span,li,div,[role='button'],[role='link'])"
    )
    seen = set()
    for i in range(min(candidates.count(), 120)):
        try:
            el = candidates.nth(i)
            if not el.is_visible():
                continue
            text = el.inner_text().strip()
            if text and re.fullmatch(r"\d+", text):
                if text not in seen:
                    seen.add(text)
                    print(f"  - {text}")
        except Exception:
            continue
    return False


def find_exact_row_on_current_page(page: Page, guild_name: str) -> Locator | None:
    exact_row: Locator | None = None
    fold_row: Locator | None = None
    for row, row_text in get_ranking_row_candidates(page):
        parsed = parse_ranking_row(row_text)
        if not parsed:
            continue
        rank, name = parsed
        print(f"row parsed: rank={rank}, name={name}")
        if name == guild_name:
            exact_row = row
            break
        if fold_row is None and name.casefold() == guild_name.casefold():
            fold_row = row
    return exact_row if exact_row is not None else fold_row


def parse_ranking_row(row_text: str) -> tuple[str, str] | None:
    compact = re.sub(r"\s+", " ", row_text).strip()
    m = re.match(r"^(?P<rank>\d+)\s+(?P<name>.+?)\s+(?P<acp>\d[\d,]*)\s+", compact)
    if not m:
        return None
    return m.group("rank"), m.group("name").strip()


def wait_for_guild_detail_transition(page: Page, guild_name: str) -> bool:
    start = time.monotonic()
    while time.monotonic() - start < 15:
        body = page.locator("body").inner_text()
        has_detail_keywords = (
            ("Guild Combat Power" in body)
            or ("Combat Power" in body)
            or ("Server Origin" in body)
            or ("Active Guild Members" in body)
        )
        guild_rank_visible = "Guild Rank" in body
        guild_name_visible = guild_name in body
        if has_detail_keywords:
            return True
        if (not guild_rank_visible) and guild_name_visible:
            return True
        time.sleep(POLL_INTERVAL_SECONDS)
    return False


def click_row_and_wait_for_detail(page: Page, row: Locator, guild_name: str) -> bool:
    name_cell = row.locator("td, [role='cell']").nth(1)
    attempts = [
        ("Guild Nameセルを click", lambda: name_cell.click(timeout=2500)),
        ("Guild Nameセルを dblclick", lambda: name_cell.dblclick(timeout=2500)),
        ("行全体を click", lambda: row.click(timeout=2500)),
        ("行全体を dblclick", lambda: row.dblclick(timeout=2500)),
        ("JS click", lambda: row.evaluate("el => el.click()")),
    ]
    for label, fn in attempts:
        try:
            print(f"{label} しました")
            fn()
            if wait_for_guild_detail_transition(page, guild_name):
                print("Guild Combat Power / Active Guild Members を確認しました")
                return True
            print("詳細ページへ遷移しなかったため次の方法を試します")
        except Exception:
            continue
    return False


def click_exact_guild_from_ranking(page: Page, guild_name: str) -> None:
    checked_rows: List[str] = []
    for page_no in [1, 2, 3, 4]:
        print(f"Guild Ranking {page_no}ページ目を検索中...")
        if page_no != 1:
            if not click_ranking_page(page, page_no):
                print(f"  - {page_no}ページ目ボタンが見つからないためスキップ")
                continue
            page.wait_for_timeout(800)

        rows = get_ranking_row_candidates(page)
        checked_rows.extend([text for _, text in rows[:30]])
        target_row = find_exact_row_on_current_page(page, guild_name)
        if target_row is not None:
            print(f"matched guild: {guild_name}")
            if click_row_and_wait_for_detail(page, target_row, guild_name):
                return

    print("Guild Ranking 候補行（1〜3ページ確認分）:")
    for text in checked_rows:
        print(f"  - {text}")
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    page.screenshot(path=str(DATA_DIR / "debug_after_search.png"), full_page=True)
    (DATA_DIR / "debug_after_search_text.txt").write_text(
        page.locator("body").inner_text(), encoding="utf-8"
    )
    raise RuntimeError(f"Guild Ranking 1〜3ページで '{guild_name}' 完全一致行をクリックできませんでした。")


def get_scroll_metrics(page: Page) -> tuple[int, int]:
    return page.evaluate("() => [Math.round(window.scrollY), Math.round(document.body.scrollHeight)]")


def wait_for_active_members_ready(page: Page) -> None:
    start = time.monotonic()
    next_log = start + 5
    while time.monotonic() - start < GUILD_LOAD_TIMEOUT_SECONDS:
        body_text = page.locator("body").inner_text()
        has_loading = "Loading Data" in body_text
        has_active_members = "Active Guild Members" in body_text
        if not has_loading and has_active_members:
            print("Active Guild Members の表示を確認しました。")
            return
        if time.monotonic() >= next_log:
            elapsed = int(time.monotonic() - start)
            print(
                f"ギルドページ読み込み待機中... ({elapsed}秒経過, "
                f"loading={'ON' if has_loading else 'OFF'}, "
                f"active_members={'ON' if has_active_members else 'OFF'})"
            )
            next_log += 5
        time.sleep(POLL_INTERVAL_SECONDS)
    raise RuntimeError("60秒以内に Active Guild Members が表示されませんでした。")


def has_member_signals(page: Page) -> bool:
    active_text = extract_active_members_text(page.locator("body").inner_text())
    if not active_text:
        return False
    return ("CPM" in active_text) or ("FCP" in active_text) or ("Lv" in active_text)


def scroll_until_member_section(page: Page) -> None:
    start = time.monotonic()
    next_log = start + 5
    while time.monotonic() - start < RESULT_TIMEOUT_SECONDS:
        y, h = get_scroll_metrics(page)
        print(f"スクロール位置: y={y}, page_height={h}")
        if has_member_signals(page):
            print("メンバー候補要素(CPM/FCP/Lv/arrow_upward)を検出しました。")
            return
        page.mouse.wheel(0, 800)
        page.wait_for_timeout(500)
        if time.monotonic() >= next_log:
            print(f"メンバー一覧待機中... ({int(time.monotonic()-start)}秒経過)")
            next_log += 5
    raise RuntimeError("30秒以内にメンバー一覧候補を検出できませんでした。")


def collect_horizontal_texts(page: Page) -> List[str]:
    script = """
() => {
  const results = [];
  const seen = new Set();
  const nodes = Array.from(document.querySelectorAll('*')).filter(el => el.scrollWidth > el.clientWidth + 20);
  for (const el of nodes) {
    const old = el.scrollLeft;
    const max = el.scrollWidth - el.clientWidth;
    const steps = [0, Math.floor(max*0.33), Math.floor(max*0.66), max];
    for (const s of steps) {
      el.scrollLeft = s;
      const txt = (el.innerText || '').trim();
      if (txt && !seen.has(txt)) { seen.add(txt); results.push(txt); }
    }
    el.scrollLeft = old;
  }
  return results;
}
"""
    return page.evaluate(script)


def dump_guild_debug_files(page: Page) -> None:
    if not DEBUG_SAVE_FILES:
        return
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    page.screenshot(path=str(DATA_DIR / "debug_guild_page.png"), full_page=True)
    body_text = page.locator("body").inner_text()
    (DATA_DIR / "debug_guild_page_text.txt").write_text(body_text, encoding="utf-8")

    blocks = page.locator(
        ":is(div,li,article,section,span,p):has-text('CPM'), "
        ":is(div,li,article,section,span,p):has-text('FCP'), "
        ":is(div,li,article,section,span,p):has-text('Lv')"
    )
    lines: List[str] = []
    for i in range(min(blocks.count(), 300)):
        t = blocks.nth(i).inner_text().strip()
        if t:
            lines.append(f"[{i+1}] {t}")
    for t in collect_horizontal_texts(page):
        if "CPM" in t or "FCP" in t or "Lv" in t:
            lines.append("[H] " + t)
    (DATA_DIR / "debug_member_blocks.txt").write_text(
        "\n\n".join(lines) if lines else "(no blocks)", encoding="utf-8"
    )


def parse_node_history_rows(body_text: str) -> List[dict[str, str]]:
    section = ""
    start = body_text.find("Node & Siege War History")
    if start != -1:
        section = body_text[start:]
        end = section.find("Guild War History")
        if end != -1:
            section = section[:end]
    if not section:
        return []

    lines = [x.strip() for x in section.splitlines() if x.strip()]
    rows: List[dict[str, str]] = []
    i = 0
    date_pattern = re.compile(r"^(\d{1,2}\s+[A-Za-z]+\s+\d{4})")
    while i < len(lines):
        m = date_pattern.match(lines[i])
        if not m:
            i += 1
            continue
        date = m.group(1)
        content_name = lines[i + 1] if i + 1 < len(lines) else ""
        result = ""
        if i + 2 < len(lines):
            rm = re.search(r"\b(Win|Lost)\b", lines[i + 2], re.I)
            if rm:
                result = rm.group(1)
        if date and content_name and result:
            rows.append({"date": date, "content_name": content_name, "result": result})
            i += 3
            continue
        i += 1
    return rows


def extract_active_members_text(text: str) -> str:
    start_marker = "Active Guild Members"
    end_marker = "Guild Member History"
    start_idx = text.find(start_marker)
    if start_idx == -1:
        return ""
    section = text[start_idx:]
    end_idx = section.find(end_marker)
    if end_idx != -1:
        section = section[:end_idx]
    return section


def parse_members_from_blocks(text: str) -> List[MemberRow]:
    retrieved_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    members: List[MemberRow] = []
    active_text = extract_active_members_text(text)
    if not active_text:
        return members

    lines = [ln for ln in active_text.splitlines() if "fmd_bad" not in ln]
    cleaned = "\n".join(lines)
    blocks = re.split(r"(?m)^\s*(?=\d+\.)", cleaned)
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        block_lines = [ln.strip() for ln in block.splitlines() if ln.strip()]
        if not block_lines:
            continue
        m_head = re.match(r"^(\d+)\.\s*(.+)$", block_lines[0])
        if not m_head:
            continue
        rank = m_head.group(1).strip()
        player_name = m_head.group(2).strip()
        level = ""
        cpm = ""
        fcp = ""
        for ln in block_lines[1:]:
            if not level:
                m = re.search(r"Lv\s*:?\s*(\d+)", ln, re.I)
                if m:
                    level = m.group(1)
            if not cpm:
                m = re.search(r"CPM\s*:?\s*([\d,]+)", ln, re.I)
                if m:
                    cpm = re.sub(r"\D", "", m.group(1))
            if fcp == "":
                m = re.search(r"FCP\s*:?\s*([\d,]+)", ln, re.I)
                if m:
                    fcp = re.sub(r"\D", "", m.group(1))
        if player_name and level and cpm and fcp != "":
            members.append(MemberRow(rank, player_name, level, cpm, fcp, retrieved_at))

    uniq_by_rank: dict[str, MemberRow] = {}
    for member in members:
        uniq_by_rank[member.rank] = member
    return list(uniq_by_rank.values())


def parse_summary_metrics(body_text: str) -> dict[str, str]:
    """ギルド概要メトリクスを body.inner_text() から抽出する。"""

    def section_between(text: str, start: str, end_markers: list[str]) -> str:
        start_idx = text.find(start)
        if start_idx == -1:
            return ""
        part = text[start_idx:]
        end_idx = len(part)
        for marker in end_markers:
            idx = part.find(marker)
            if idx != -1:
                end_idx = min(end_idx, idx)
        return part[:end_idx]

    def norm_num(v: str) -> str:
        return re.sub(r"[^0-9]", "", v)

    combat = section_between(body_text, "Combat Power", ["Guild War Activity", "Node War Activity"])
    guild_war = section_between(
        body_text, "Guild War Activity", ["Node War Activity", "Node & Siege War History"]
    )
    node_war = section_between(
        body_text,
        "Node War Activity",
        ["Known Guild Names", "CP Distributions", "Node & Siege War History"],
    )

    def pick_after_label(section: str, label: str) -> str:
        m_inline = re.search(re.escape(label) + r"\s*:?\s*([^\n]+)", section, re.I)
        if m_inline:
            value = m_inline.group(1).strip()
            if value and value.lower() != label.lower():
                return value
        lines = [ln.strip() for ln in section.splitlines() if ln.strip()]
        for i, ln in enumerate(lines):
            if re.fullmatch(re.escape(label), ln, re.I) and i + 1 < len(lines):
                return lines[i + 1]
        return ""

    low_member_cp = ""
    high_member_cp = ""
    lines = [ln.strip() for ln in combat.splitlines() if ln.strip()]
    for i, ln in enumerate(lines):
        if re.fullmatch(r"ACTIVE LOW & HIGH POINT MEMBER CP", ln, re.I):
            look = lines[i + 1 : i + 8]
            for v in look:
                m = re.search(r"([\d,]+)\s*\((L|H)\)", v, re.I)
                if not m:
                    continue
                if m.group(2).upper() == "L":
                    low_member_cp = norm_num(m.group(1))
                else:
                    high_member_cp = norm_num(m.group(1))
            break

    return {
        "avg_cp": norm_num(pick_after_label(combat, "AVERAGE CP FROM ACTIVE MEMBER")),
        "total_cp": norm_num(pick_after_label(combat, "TOTAL CP FROM ACTIVE MEMBER")),
        "total_family_cp": norm_num(
            pick_after_label(combat, "TOTAL FAMILY CP FROM ACTIVE MEMBER")
        ),
        "active_member_count": norm_num(
            pick_after_label(combat, "ACTIVE MEMBER - EXCLUDE MEMBERS WHO MOSTLY BSM")
        ),
        "low_member_cp": low_member_cp,
        "high_member_cp": high_member_cp,
        "declared_on_other_guild": norm_num(
            pick_after_label(guild_war, "DECLARE ON OTHER GUILD")
        ),
        "declared_by_other_guild": norm_num(
            pick_after_label(guild_war, "DECLARED BY OTHER GUILD")
        ),
        "total_war": norm_num(pick_after_label(guild_war, "TOTAL WAR")),
        "all_time_win_rate": pick_after_label(guild_war, "ALL TIME WIN RATE").strip(),
        "most_war_with_guild": pick_after_label(guild_war, "MOST WAR WITH GUILD").strip(),
        "total_node_wars": norm_num(pick_after_label(node_war, "TOTAL NODE WARS")),
        "node_won": norm_num(pick_after_label(node_war, "NODE WON")),
        "total_siege_wars": norm_num(pick_after_label(node_war, "TOTAL SIEGE WARS")),
        "siege_won": norm_num(pick_after_label(node_war, "SIEGE WON")),
        "currently_holding": pick_after_label(node_war, "CURRENTLY HOLDING").strip(),
    }


def save_summary_to_csv(guild_name: str, body_text: str) -> Path:
    today = datetime.now().strftime("%Y-%m-%d")
    retrieved_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    safe_guild_name = sanitize_filename(guild_name)
    file_path = get_guild_data_dir(guild_name) / f"summary_{safe_guild_name}_{today}.csv"
    metrics = parse_summary_metrics(body_text)
    row = {"guild_name": guild_name, "retrieved_at": retrieved_at, **metrics}
    cols = get_summary_columns()
    with file_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        w.writerow(row)
    return file_path


def save_node_history_to_csv(guild_name: str, body_text: str) -> Path:
    today = datetime.now().strftime("%Y-%m-%d")
    retrieved_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    safe_guild_name = sanitize_filename(guild_name)
    file_path = get_guild_data_dir(guild_name) / f"node_history_{safe_guild_name}_{today}.csv"
    rows = [
        {
            "guild_name": guild_name,
            "retrieved_at": retrieved_at,
            "date": row["date"],
            "content_name": row["content_name"],
            "result": row["result"],
        }
        for row in parse_node_history_rows(body_text)
    ]
    cols = ["guild_name", "retrieved_at", "date", "content_name", "result"]
    with file_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols)
        w.writeheader()
        for row in rows:
            w.writerow(row)
    return file_path


def get_summary_columns() -> list[str]:
    return [
        "guild_name",
        "retrieved_at",
        "avg_cp",
        "total_cp",
        "total_family_cp",
        "active_member_count",
        "low_member_cp",
        "high_member_cp",
        "declared_on_other_guild",
        "declared_by_other_guild",
        "total_war",
        "all_time_win_rate",
        "most_war_with_guild",
        "total_node_wars",
        "node_won",
        "total_siege_wars",
        "siege_won",
        "currently_holding",
    ]


def save_guild_workbook(guild_name: str, members: List[MemberRow], body_text: str) -> Path:
    from openpyxl import Workbook

    today = datetime.now().strftime("%Y-%m-%d")
    retrieved_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    safe_guild_name = sanitize_filename(guild_name)
    guild_dir = get_guild_data_dir(guild_name)
    file_path = guild_dir / f"guild_{safe_guild_name}_{today}.xlsx"

    wb = Workbook()
    ws_members = wb.active
    ws_members.title = "members"
    ws_members.append(["rank", "player_name", "level", "cpm", "fcp", "retrieved_at"])
    for m in members:
        ws_members.append([m.rank, m.player_name, m.level, m.cpm, m.fcp, m.retrieved_at])

    ws_summary = wb.create_sheet("summary")
    summary_cols = get_summary_columns()
    ws_summary.append(summary_cols)
    metrics = parse_summary_metrics(body_text)
    ws_summary.append([guild_name, retrieved_at] + [metrics.get(c, "") for c in summary_cols[2:]])

    ws_node = wb.create_sheet("node_history")
    ws_node.append(["guild_name", "retrieved_at", "date", "content_name", "result"])
    for row in parse_node_history_rows(body_text):
        ws_node.append([guild_name, retrieved_at, row["date"], row["content_name"], row["result"]])

    wb.save(file_path)
    return file_path


def save_guild_snapshot_to_sqlite(
    guild_name: str, members: List[MemberRow], body_text: str
) -> bool:
    """Save scraper results to SQLite without affecting Excel success."""

    retrieved_at = members[0].retrieved_at if members else datetime.now().strftime("%Y-%m-%d %H:%M")
    member_rows = [
        {
            "rank_no": m.rank,
            "family_name": m.player_name,
            "level": m.level,
            "cpm": m.cpm,
            "fcp": m.fcp,
            "class_name": None,
            "class_name_raw": None,
            "class_name_normalized": None,
            "class_name_version": None,
        }
        for m in members
    ]
    node_history = [
        {
            "row_no": index,
            "war_date": row.get("date", ""),
            "node_name": row.get("content_name", ""),
            "opponent_guild": "",
            "result": row.get("result", ""),
            "raw": row,
        }
        for index, row in enumerate(parse_node_history_rows(body_text), start=1)
    ]
    summary = parse_summary_metrics(body_text)
    try:
        conn = db.initialize()
        try:
            db.save_snapshot(
                conn,
                retrieved_at=retrieved_at,
                guild_name=guild_name,
                members=member_rows,
                member_count=len(member_rows),
                node_history=node_history,
                summary=summary,
            )
        finally:
            conn.close()
    except Exception as exc:
        print(f"⚠ SQLite保存に失敗しました（Excelは作成済みです）: {exc}")
        return False
    print(f"SQLite保存完了: {db.DEFAULT_DB_PATH}")
    return True


def save_scraped_data_to_sqlite(
    *,
    retrieved_at: str,
    guild_name: str,
    members: list[dict[str, object]],
    db_path: str | Path | None = None,
    member_count: int | None = None,
    avg_cpm: float | None = None,
    total_cpm: float | None = None,
    node_history: list[dict[str, object]] | None = None,
    summary: dict[str, object] | None = None,
) -> bool:
    """Compatibility helper for callers that already have dict rows."""

    try:
        conn = db.initialize(db_path)
        try:
            db.save_snapshot(
                conn,
                retrieved_at=retrieved_at,
                guild_name=guild_name,
                members=members,
                member_count=member_count,
                avg_cpm=avg_cpm,
                total_cpm=total_cpm,
                node_history=node_history,
                summary=summary,
            )
        finally:
            conn.close()
    except Exception as exc:
        print(f"⚠ SQLite保存に失敗しました（Excelは作成済みです）: {exc}")
        return False
    return True


def parse_expected_active_member_count(text: str) -> int | None:
    patterns = [
        r"ACTIVE MEMBER - EXCLUDE MEMBERS WHO MOSTLY BSM\s*:?\s*(\d+)",
        r"ACTIVE MEMBER\s*:?\s*(\d+)",
    ]
    for pattern in patterns:
        m = re.search(pattern, text, flags=re.IGNORECASE)
        if m:
            return int(m.group(1))
    return None


def scrape_members_from_guild_page(page: Page) -> List[MemberRow]:
    wait_for_active_members_ready(page)
    scroll_until_member_section(page)
    y, h = get_scroll_metrics(page)
    print(f"取得前スクロール位置: y={y}, page_height={h}")
    dump_guild_debug_files(page)
    body_text = page.locator("body").inner_text()
    active_text = extract_active_members_text(body_text)
    if not active_text:
        print("Active Guild Members セクションの切り出しに失敗しました。")
    members = parse_members_from_blocks(body_text)
    if not members:
        raise RuntimeError("ギルドページでメンバー情報を取得できませんでした。debug_member_blocks.txt を確認してください。")
    return members


def save_members_to_csv(guild_name: str, members: List[MemberRow]) -> Path:
    today = datetime.now().strftime("%Y-%m-%d")
    safe_guild_name = sanitize_filename(guild_name)
    file_path = get_guild_data_dir(guild_name) / f"members_{safe_guild_name}_{today}.csv"
    with file_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["rank", "player_name", "level", "cpm", "fcp", "retrieved_at"])
        for m in members:
            w.writerow([m.rank, m.player_name, m.level, m.cpm, m.fcp, m.retrieved_at])
    return file_path


def return_to_guild_ranking(page: Page) -> None:
    back_candidates = [
        page.locator("button:has-text('arrow_back')"),
        page.locator("[aria-label*='back' i]"),
        page.locator("button i:has-text('arrow_back')").locator("xpath=ancestor::button[1]"),
    ]
    for c in back_candidates:
        try:
            if c.count() > 0 and c.first.is_visible():
                c.first.click(timeout=2000)
                page.wait_for_timeout(500)
                break
        except Exception:
            continue
    try:
        open_guild_ranking_page(page)
    except Exception:
        pass
    start = time.monotonic()
    while time.monotonic() - start < RESULT_TIMEOUT_SECONDS:
        try:
            if get_guild_ranking_search_input(page).is_visible():
                return
        except Exception:
            pass
        time.sleep(POLL_INTERVAL_SECONDS)
    raise RuntimeError("Guild Ranking 画面へ戻れませんでした。")


def search_and_open_guild(page: Page, guild_name: str) -> None:
    close_general_chat_panel(page)
    open_guild_ranking_page(page)
    search_guild_in_ranking(page, guild_name)
    click_exact_guild_from_ranking(page, guild_name)


def main() -> int:
    from dotenv import load_dotenv
    from playwright.sync_api import (
        TimeoutError as PlaywrightTimeoutError,
        sync_playwright,
    )

    load_dotenv()

    try:
        guilds = load_guild_names(CONFIG_PATH)
    except Exception as e:
        print(f"❌ 初期化エラー: {e}")
        return 1

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        ctx = browser.new_context()
        page = ctx.new_page()
        page.goto(DBONK_LOGIN_URL, wait_until="domcontentloaded")
        try_auto_login(page)
        try_select_asia_server(page)
        page.wait_for_timeout(1000)
        total, ok = len(guilds), 0
        try:
            for i, guild in enumerate(guilds, 1):
                print(f"[{i}/{total}] {guild} ... 取得中")
                try:
                    search_and_open_guild(page, guild)
                    members = scrape_members_from_guild_page(page)
                    body_text = page.locator("body").inner_text()
                    workbook_path = save_guild_workbook(guild, members, body_text)
                    save_guild_snapshot_to_sqlite(guild, members, body_text)
                    print(f"取得メンバー数: {len(members)}")
                    expected_count = parse_expected_active_member_count(body_text)
                    if expected_count is not None and expected_count != len(members):
                        print(
                            f"⚠ 表示上のActive Memberは{expected_count}人ですが、"
                            f"CSV取得は{len(members)}人です"
                        )
                    ok += 1
                    print(f"[{i}/{total}] {guild} ... 取得完了 -> {workbook_path}")
                    if i < total:
                        return_to_guild_ranking(page)
                except (PlaywrightTimeoutError, RuntimeError) as e:
                    print(f"[{i}/{total}] {guild} ... エラー: {e}")
                except Exception as e:
                    print(f"[{i}/{total}] {guild} ... 想定外エラー: {e}")
            print(f"✅ 全処理終了: {ok}/{total} ギルドの取得が完了しました。")
            return 0
        finally:
            input("確認したらEnterで終了")
            ctx.close()
            browser.close()


if __name__ == "__main__":
    sys.exit(main())
